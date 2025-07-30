import csv
import io
from typing import Any, cast

from openpyxl import Workbook


def _get_nesting_cols_by_name(nesting_keys: list[str], fieldnames: list[str]) -> list[str]:
    """Get nesting columns by name."""
    actual_nesting_cols = []
    fieldnames_lower = {f.lower(): f for f in fieldnames}
    for key in nesting_keys:
        actual_col = fieldnames_lower.get(str(key).lower())
        if not actual_col:
            raise ValueError(f"CSV must contain column '{key}'.")
        actual_nesting_cols.append(actual_col)
    return actual_nesting_cols


def _get_nesting_cols_by_index(nesting_keys: list[int], fieldnames: list[str]) -> list[str]:
    """Get nesting columns by index."""
    actual_nesting_cols = []
    for index in nesting_keys:
        try:
            actual_nesting_cols.append(fieldnames[int(index)])
        except IndexError as e:
            raise IndexError(f"Column index {index} is out of range.") from e
    return actual_nesting_cols


def _get_actual_nesting_cols(nesting_keys: list[str] | list[int], fieldnames: list[str]) -> list[str]:
    """Resolves nesting keys to actual column names from fieldnames."""
    if not nesting_keys:
        return []

    if isinstance(nesting_keys[0], str):
        return _get_nesting_cols_by_name(cast(list[str], nesting_keys), fieldnames)
    if isinstance(nesting_keys[0], int):
        return _get_nesting_cols_by_index(cast(list[int], nesting_keys), fieldnames)

    raise TypeError("nesting_keys must be a list of strings or a list of integers.")


def _build_nested_structure(reader: csv.DictReader, actual_nesting_cols: list[str]) -> dict:
    """Builds the nested dictionary from the CSV reader."""
    result = {}
    for row in reader:
        current_level = result
        # Iterate through nesting keys to build the nested structure
        for i, col_name in enumerate(actual_nesting_cols):
            key_value = row.pop(col_name)

            if i < len(actual_nesting_cols) - 1:
                # This is not the last nesting level, so ensure a dict exists
                if key_value not in current_level:
                    current_level[key_value] = {}
                current_level = current_level[key_value]
            else:
                # This is the last nesting level, assign the rest of the row
                current_level[key_value] = row
    return result


def csv_to_nested_dict(csv_string: str, nesting_keys: list[str] | list[int]) -> dict:
    """
    Converts a CSV string to a nested dictionary based on a specified list of nesting keys.

    The nesting keys determine the structure of the output dictionary.
    Keys can be column names (str) or column indices (int).
    Column name matching for nesting keys is case-insensitive.
    """
    f = io.StringIO(csv_string.strip())
    reader = csv.DictReader(f)

    if not reader.fieldnames:
        return {}

    actual_nesting_cols = _get_actual_nesting_cols(nesting_keys, list(reader.fieldnames))
    if not actual_nesting_cols:
        return {}

    return _build_nested_structure(reader, actual_nesting_cols)


def list_dict_to_csv(data: list[dict]) -> str:
    """
    Converts a list of dictionaries to a CSV formatted string.

    The keys of the first dictionary in the list are used as the header row.
    Handles values with commas and other special characters by quoting them.

    Args:
        data: A list of dictionaries. Each dictionary represents a row.
              It's assumed that all dictionaries have the same keys.

    Returns:
        A string containing the data in CSV format.
        Returns an empty string if the input list is empty.
    """
    if not data:
        return ""

    # Use an in-memory string buffer to build the CSV string.
    # The `newline=''` argument is important to prevent extra blank rows.
    output = io.StringIO()

    # The fieldnames are the keys from the first dictionary.
    fieldnames = data[0].keys()

    # Create a DictWriter object. This object can map dictionaries to CSV rows.
    # The `quoting=csv.QUOTE_MINIMAL` (the default) ensures that fields are only
    # quoted if they contain the delimiter, quotechar, or any of the characters
    # in lineterminator.
    writer = csv.DictWriter(output, fieldnames=fieldnames)

    # Write the header row (the dictionary keys).
    writer.writeheader()

    # Write all the dictionary data to the CSV buffer.
    writer.writerows(data)

    # Return the complete CSV string from the buffer.
    return output.getvalue()


def list_dict_to_xlsx_bytes(data: list[dict[str, Any]]) -> bytes:
    """
    Converts a list of dictionaries to an XLSX file as bytes.

    Args:
        data: A list of dictionaries. Each dictionary represents a row.
              It's assumed that all dictionaries have the same keys.

    Returns:
        The XLSX file content as bytes.
        Returns empty bytes if the input list is empty.
    """
    if not data:
        return b""

    workbook = Workbook()
    sheet = workbook.active

    # Write headers
    headers = list(data[0].keys())
    for col_num, header_title in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header_title

    # Write data rows
    for row_num, row_data in enumerate(data, 2):
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = row_data.get(header)

    # Save to a bytes buffer
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
