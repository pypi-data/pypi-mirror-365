"""ExcelReader class for extracting data from Excel files."""
import warnings
import zipfile
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.utils.exceptions import InvalidFileException


class ExcelReader:
    """
    A class to read and extract data from Excel files (.xlsx, .xlsm, .xlsb).

    This class provides methods to load workbooks, extract data from sheets,
    tables, and named ranges, with robust error handling.
    """

    def __init__(self, file_path: Path | str):
        """
        Initializes the ExcelReader with the path to the Excel file.

        Args:
            file_path (str | Path): Path to the Excel file, specified as a Path object or string.

        Raises:
            ImportError: If the file does not exist, is not a valid Excel file, or if the openpyxl library cannot be imported.
        """
        if isinstance(file_path, str):
            file_path = Path(file_path)
        if not file_path.is_file():
            msg = f"File {file_path} does not exist or is not a valid file."
            raise ImportError(msg)

        # Check extension to see if it's an Excel file
        if file_path.suffix.lower() not in {".xlsx", ".xlsm", ".xlsb"}:
            msg = f"File {file_path} is not a valid Excel file."
            raise ImportError(msg)

        self.file_path = file_path

    def load_excel_workbook(self, data_only=True, read_only=False):  # noqa: FBT002
        """
        Load an Excel workbook with robust error handling.

        Args:
            data_only (Optional[bool], optional): Whether to return cell values (not formulas).
            read_only (Optional[bool], optional): Use openpyxl's read-only mode for large files.

        Raises:
            ImportError: If the file cannot be loaded due to various reasons (e.g., file not found, permission denied, invalid format).

        Returns:
            object (workbook): Workbook object or None if loading fails.
        """
        warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
        try:
            wb = load_workbook(filename=self.file_path, data_only=data_only, read_only=read_only)
        except FileNotFoundError:
            msg = f"File not found: {self.file_path}"
        except PermissionError:
            msg = f"Permission denied: {self.file_path}"
        except InvalidFileException:
            msg = f"Invalid Excel file format: {self.file_path}"
        except zipfile.BadZipFile:
            msg = f"Corrupt Excel file (not a valid ZIP): {self.file_path}"
        except OSError as e:
            msg = f"Unexpected error loading Excel file {self.file_path}: {e}"
        else:
            if wb is None:
                msg = f"Failed to load workbook from {self.file_path}. The file may be corrupted or not a valid Excel file."
            else:
                return wb

        raise ImportError(msg)   # Raise an error with the message if loading fails

    def extract_data(self, source_name: str, source_type: str) -> list[dict]:
        """
        Extracts data from an Excel file based on the source type and name.

        Expected the specified source type to be either:
        - An entire worksheet with the header in the first row (sheet)
        - A named range (range)
        - An Excel table (table)

        Args:
            source_name (str): Name of the sheet, table, or range to extract.
            source_type (str): Type of source ('sheet', 'table', or 'range').

        Raises:
            ImportError: If the source type is invalid or if there are issues extracting data.

        Returns:
            data (list[dict]): Data extracted as a dictionary.
        """
        if source_type == "sheet":
            return self.extract_from_sheet(source_name)
        if source_type == "table":
            return self.extract_from_table(source_name)
        if source_type == "range":
            return self.extract_from_range(source_name)

        msg = f"Invalid source type '{source_type}'. Must be 'sheet', 'table', or 'range'."
        raise ImportError(msg)

    def extract_from_sheet(self, sheet_name: str) -> list[dict]:
        """
        Extracts a sheet from an Excel file and returns it as a DataFrame.

        Args:
            sheet_name (str): Name of the sheet to extract.

        Raises:
            ImportError: If the sheet cannot be loaded or if there are issues with the file.

        Returns:
            data (list[dict]): A list containing the sheet data.
        """
        try:
            table_df = pd.read_excel(self.file_path, sheet_name=sheet_name)
            table_df.columns = table_df.columns.str.strip()
            table_data = table_df.to_dict(orient="records")
        except ImportError as e:
            msg = f"Error importing the openpyxl library for reading Excel file {self.file_path}: {e}. "
            msg += "Try installing it with 'pip install openpyxl' or 'uv add openpyxl'."
        except zipfile.BadZipFile:
            msg = f"Corrupt Excel file (not a valid ZIP): {self.file_path}"
        except OSError as e:
            msg = f"Unexpected error loading Excel file {self.file_path}: {e}"
        except (ValueError, AttributeError) as e:
            msg = f"Error loading sheet '{sheet_name}' from Excel file {self.file_path}: {e}"
        else:
            return table_data

        # If we reach here, it means the import failed
        raise ImportError(msg)

    def extract_from_table(self, table_name: str) -> list[dict]:
        """
        Extracts a table from an Excel file and returns it as a DataFrame.

        Args:
            table_name (str): Name of the table to extract.

        Raises:
            ImportError: If the table cannot be loaded or if there are issues with the file.

        Returns:
            data (list[dict]): A list containing the table data.
        """
        # load_excel_workbook() will raise an ImportError if there's an issue. Let this be caught by the caller.
        wb = self.load_excel_workbook(data_only=True, read_only=False)
        found = False
        selected_sheet = None
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if table_name in ws.tables:
                table = ws.tables[table_name]
                table_range = table.ref  # e.g., "B3:F12"
                min_col, min_row, max_col, max_row = range_boundaries(table_range)
                found = True
                selected_sheet = sheet_name
                break  # Stop scanning after finding the first match

        if not found:
            msg = f"Table '{table_name}' not found in any worksheet of the Excel file {self.file_path}."
            raise ImportError(msg)

        if min_col is None or max_col is None or max_row is None or min_row is None:
            msg = f"Invalid table range boundaries for table '{table_name}' in file {self.file_path}."
            raise ImportError(msg)

        try:
            # Generate usecols as Excel letters (e.g., 'B', 'C', ..., 'F')
            usecols = [get_column_letter(col) for col in range(min_col, max_col + 1)]
            # Include all rows in the table
            nrows = max_row - min_row + 1  # type: ignore[assignment]

            # Read the specific table range using pandas
            table_df = pd.read_excel(
                self.file_path,
                sheet_name=selected_sheet,
                usecols=",".join(usecols),
                skiprows=min_row - 1,   # skip up to the header row (Excel is 1-based)
                nrows=nrows,
                header=None     # Don't treat the first row as header so that we avoid the bug where pandas renames columns it thinks are duplicates
            )

            # Strip whitespace from column names
            table_df.columns = table_df.iloc[0].map(str).str.strip()  # type: ignore[attr-defined]

            # Define the first row as the header
            table_df = table_df[1:].reset_index(drop=True)  # type: ignore[attr-defined]

            # Extract to dictionary
            table_data = table_df.to_dict(orient="records")

        except (ImportError, ValueError, AttributeError) as e:
            msg = f"Error loading table '{table_name}' from Excel file {self.file_path}: {e}"
            raise ImportError(msg) from e
        else:
            return table_data

    def extract_from_range(self, range_name: str) -> list[dict]:
        """
        Extracts a table from an Excel file and returns it as a DataFrame.

        Args:
            range_name (str): Name of the range to extract.

        Raises:
            ImportError: If the table cannot be loaded or if there are issues with the file.

        Returns:
            data (list[dict]): A list containing the range data.
        """
        # load_excel_workbook() will raise an ImportError if there's an issue. Let this be caught by the caller.
        wb = self.load_excel_workbook(data_only=True, read_only=True)
        found = False

        for defined_name in wb.defined_names:
            if defined_name == range_name:
                dn = wb.defined_names[defined_name]
                destinations = list(dn.destinations)
                if len(destinations) != 1:
                    msg = f"Range '{range_name}' refers to multiple areas, which is unsupported."
                    raise ImportError(msg)

                sheet_name, ref = destinations[0]  # unpack tuple
                min_col, min_row, max_col, max_row = range_boundaries(ref)
                found = True
                break

        if not found:
            msg = f"Range '{range_name}' not found in any worksheet of the Excel file {self.file_path}."
            raise ImportError(msg)

        if min_col is None or max_col is None or max_row is None or min_row is None:
            msg = f"Invalid table range boundaries for range '{range_name}' in file {self.file_path}."
            raise ImportError(msg)

        try:
            # Generate usecols as Excel letters (e.g., 'B', 'C', ..., 'F')
            usecols = [get_column_letter(col) for col in range(min_col, max_col + 1)]
            # Include all rows in the table
            nrows = max_row - min_row + 1    # type: ignore[assignment]

            # Read the specific table range using pandas
            table_df = pd.read_excel(
                self.file_path,
                sheet_name=sheet_name,
                usecols=",".join(usecols),
                skiprows=min_row - 1,  # skip to the first row of the range
                nrows=nrows,
                header=None     # Don't treat the first row as header so that we avoid the bug where pandas renames columns it thinks are duplicates
            )

            # Strip whitespace from column names
            table_df.columns = table_df.iloc[0].map(str).str.strip()  # type: ignore[attr-defined]

            # Define the first row as the header
            table_df = table_df[1:].reset_index(drop=True)

            # Extract to dictionary
            table_data = table_df.to_dict(orient="records")

        except (ImportError, ValueError, AttributeError) as e:
            msg = f"Error loading data range '{range_name}' from Excel file {self.file_path}: {e}"
            raise ImportError(msg) from e
        else:
            return table_data


