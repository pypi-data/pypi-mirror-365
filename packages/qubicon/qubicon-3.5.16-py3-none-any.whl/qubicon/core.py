# CUSTOM SDK LOGIC (ONTOP OF AUTO-GENERATED OPENAPI CLIENT)

import time
from qubicon.api.public_api_controller import (
    create_computable_model,
    create_physical_quantity,
    get_channels,
    get_equipment_variable_nodes,
    get_job_artifact,
    get_job,
    get_list_of_computable_models,
    get_multiplex_chart_data_channels,
    get_offline_equipment,
    get_online_equipments,
    get_physical_quantities,
    get_processes,
    start_export_process_data,
    stream_events,
    login,
    refresh,
)

from qubicon.api.client import AuthenticatedClient
from qubicon.api.types import UNSET, Unset

from qubicon.models import (
    AbstractComputableModelDto,
    AbstractComputableModelDtoEngineType,
    AbstractComputableModelDtoCalculationStyle,
    AbstractComputableModelDtoStatus,
    ChannelDataChannelRequestKeyDto,
    ChannelDataChannelRequestKeyDtoType,
    ChannelDataPairDto,
    ComputableModelInputDto,
    ComputableModelOutputDto,
    ExternalPythonComputableModelDto,
    JythonComputableModelDto,
    MultiplexChartDataChannelsRequestDto,
    MultiplexChartDataDto,
    PhysicalQuantityDto,
    PhysicalQuantityUnitDto,
    PhysicalQuantityUnitDtoStatus,
    PublicProcessExportRequestDto,
    RefreshTokenDto,
    StoredFileDto,
    UserLoginDto,
    UserLoginDtoLoginType,
)
import pandas as pd
from http import HTTPStatus
from typing import Any, Dict, List, Optional
import json
import logging
from rapidfuzz import fuzz
import re
import os

from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import requests


class QubiconCore:
    def __init__(self, client: AuthenticatedClient):
        """
        Keep the entire HTTP client here, which contains:
            - base_url
            - token (updated upon login)
            - any request config
        """
        self.client = client

    def login_user(
        self, username: str, password: str, login_type: str = "INTERNAL"
    ) -> Optional[str]:
        """Login user and return authentication token."""
        # Remove any stale token header (only for this login call)
        httpx_client = self.client.get_httpx_client()
        httpx_client.headers.pop("Authorization", None)

        login_type_enum = UserLoginDtoLoginType(login_type)
        body = UserLoginDto(
            username=username, password=password, login_type=login_type_enum
        )

        try:
            response = login.sync_detailed(client=self.client, body=body)

            if response.status_code == HTTPStatus.OK:
                parsed_response = json.loads(response.content.decode())

                token = parsed_response.get("normal", {}).get("token")
                if not token:
                    raise ValueError("Token not found in response.")

                # Update the client with the new token
                self.client.token = token

                # Re-add the Authorization header for future requests.
                httpx_client.headers["Authorization"] = f"Bearer {token}"

                logging.debug("Login successful! Token set in client.")
                return token
            else:
                logging.error(f"Login failed with status code: {response.status_code}")
                return None
        except Exception as e:
            logging.error(f"Error logging in: {e}")
            return None

    def login_user1(
        self, username: str, password: str, login_type: str = "INTERNAL"
    ) -> Optional[str]:
        """Login user and return authentication token."""
        login_type_enum = UserLoginDtoLoginType(login_type)
        body = UserLoginDto(
            username=username, password=password, login_type=login_type_enum
        )
        try:
            response = login.sync_detailed(client=self.client, body=body)
            if response.status_code == HTTPStatus.OK:
                print("Login successful!")
                return response.parsed.get("token")
            else:
                print(f"Login failed: {response.content}")
                return None
        except Exception as e:
            print(f"Error logging in: {e}")
            return None

    def refresh_token(self, refresh_token: str) -> Optional[str]:
        """Refresh authentication token."""
        body = RefreshTokenDto(refresh_token=refresh_token)
        try:
            response = refresh.sync_detailed(client=self.client, body=body)
            if response.status_code == HTTPStatus.OK:
                print("Token refreshed successfully!")
                return response.parsed.get("accessToken")
            else:
                print(f"Token refresh failed: {response.content}")
                return None
        except Exception as e:
            print(f"Error refreshing token: {e}")
            return None

    def list_computable_models(
        self,
        statuses: list = UNSET,
        calculation_styles: list = UNSET,
        types: list = UNSET,
        output_physical_quantity_unit_ids: list = UNSET,
        output_physical_quantity_ids: list = UNSET,
        name: str = UNSET,
        deleted: bool = UNSET,
        sort: list = UNSET,
        page: int = UNSET,
        size: int = 1000,
    ):
        """
        Fetch all computable models with optional filters applied.

        Args:
            statuses (list, optional): Filter by statuses (e.g., ['RELEASED', 'DRAFT']).
            calculation_styles (list, optional): Filter by calculation styles.
            types (list, optional): Filter by model types.
            output_physical_quantity_unit_ids (list, optional): Filter by output physical quantity unit IDs.
            output_physical_quantity_ids (list, optional): Filter by output physical quantity IDs.
            name (str, optional): Filter by name.
            deleted (bool, optional): Filter by deletion status.
            sort (list, optional): Sort results (e.g., ['name,asc']).
            page (int, optional): Specify the page number for pagination.
            size (int, optional): Specify the number of results per page (default: 1000).

        Returns:
            List[Dict] or Dict: A list of computable models or an error message in JSON format.
        """
        try:
            # Fetch computable models using the OpenAPI function with filters applied
            response = get_list_of_computable_models.sync_detailed(
                client=self.client,
                statuses=statuses,
                calculation_styles=calculation_styles,
                types=types,
                output_physical_quantity_unit_ids=output_physical_quantity_unit_ids,
                output_physical_quantity_ids=output_physical_quantity_ids,
                name=name,
                deleted=deleted,
                sort=sort,
                page=page,
                size=size,
            )

            # Ensure the response status is OK
            if response.status_code == HTTPStatus.OK:
                # Parse the JSON content
                return json.loads(response.content.decode())
            else:
                logging.error(
                    f"Failed to fetch computable models. HTTP Status: {response.status_code}"
                )
                return {
                    "error": f"Failed to fetch computable models. HTTP Status: {response.status_code}"
                }

        except Exception as e:
            logging.error(f"Error fetching computable models: {e}")
            return {"error": f"Error fetching computable models: {str(e)}"}

    def fetch_model_details(self, model_id: int):
        """
        Fetch details for a specific model by filtering from the list of models.

        Args:
            model_id (int): The ID of the model to fetch details for.

        Returns:
            Dict or None: The details of the model if found, or None if not found.
        """
        try:
            # Fetch all computable models
            response = get_list_of_computable_models.sync_detailed(
                client=self.client, size=1000
            )

            if response.status_code == HTTPStatus.OK:
                # Parse the JSON content
                parsed_response = json.loads(response.content.decode())

                # Check if 'content' exists and filter for the model with the given ID
                if "content" in parsed_response and isinstance(
                    parsed_response["content"], list
                ):
                    models = parsed_response["content"]

                    # Search for the model with the matching ID
                    for model in models:
                        if model["id"] == model_id:
                            return model  # Return the specific model

                    # If the ID is not found, return None
                    logging.warning(f"Model with ID {model_id} not found.")
                    return None
                else:
                    logging.error(
                        "Unexpected response format: 'content' key not found or invalid."
                    )
                    return None
            else:
                logging.error(
                    f"Failed to fetch models. HTTP Status: {response.status_code}"
                )
                return None

        except Exception as e:
            logging.error(f"Error fetching model details: {e}")
            return None

        except Exception as e:
            logging.error(f"Error fetching model details: {e}")
            return None

    def list_physical_quantities(self):
        """
        List all physical quantities and return them as JSON.

        Returns:
            List[Dict] or Dict: A list of physical quantities as dictionaries or an error message if the request fails.
        """
        try:
            # Fetch the physical quantities using the OpenAPI function
            response = get_physical_quantities.sync_detailed(client=self.client)

            # Ensure the response status is OK
            if response.status_code == HTTPStatus.OK:
                # Parse and return the JSON content
                return json.loads(response.content.decode())
            else:
                # Log and return an error message in JSON format
                logging.error(
                    f"Failed to fetch physical quantities. HTTP Status: {response.status_code}"
                )
                logging.error(f"Response: {response.content.decode()}")
                return {
                    "error": f"Failed to fetch physical quantities. HTTP Status: {response.status_code}"
                }

        except Exception as e:
            # Log and return an error message in JSON format
            logging.error(f"Error fetching physical quantities: {e}")
            return {"error": f"Error fetching physical quantities: {str(e)}"}

    def list_processes(
        self,
        name: str = UNSET,
        deleted: bool = UNSET,
        start_date: int = UNSET,
        end_date: int = UNSET,
        statuses: list = UNSET,
        archived_or_will_be_archived: bool = UNSET,
        types: list = UNSET,
        ids: list = UNSET,
        lot_ids: list = UNSET,
        organism_ids: list = UNSET,
        kpi_model_ids: list = UNSET,
        material_ids: list = UNSET,
        stream_ids: list = UNSET,
        user_stream_ids: list = UNSET,
        online_equipment_ids: list = UNSET,
        offline_equipment_ids: list = UNSET,
        online_equipment_group_first_ids: list = UNSET,
        offline_equipment_group_first_ids: list = UNSET,
        recipe_ids: list = UNSET,
        organism_and_vial_tuples: list = UNSET,
        material_and_lot_tuples: list = UNSET,
        gmp: bool = UNSET,
        imported: bool = UNSET,
        group_ids: list = UNSET,
        master_recipe_ids: list = UNSET,
        part_of_experiment: bool = UNSET,
        sampling_values: list = UNSET,
        sort: list = UNSET,
        page: int = UNSET,
        size: int = 10000,
    ):
        """
        List all processes with optional filters applied.

        Args:
            name (str, optional): Filter by process name.
            deleted (bool, optional): Filter by deleted status.
            start_date (int, optional): Filter by processes starting on/after this date (Unix timestamp in ms).
            end_date (int, optional): Filter by processes ending on/before this date (Unix timestamp in ms).
            statuses (list, optional): Filter by process statuses (e.g., ['COMPLETED']).
            archived_or_will_be_archived (bool, optional): Filter by archived or to-be-archived processes.
            types (list, optional): Filter by process types (e.g., ['IMPORT']).
            ids (list, optional): Filter by specific process IDs.
            lot_ids (list, optional): Filter by specific lot IDs.
            organism_ids (list, optional): Filter by specific organism IDs.
            kpi_model_ids (list, optional): Filter by specific KPI model IDs.
            material_ids (list, optional): Filter by specific material IDs.
            stream_ids (list, optional): Filter by specific stream IDs.
            user_stream_ids (list, optional): Filter by specific user-defined stream IDs.
            online_equipment_ids (list, optional): Filter by specific online equipment IDs.
            offline_equipment_ids (list, optional): Filter by specific offline equipment IDs.
            online_equipment_group_first_ids (list, optional): Filter by first online equipment group IDs.
            offline_equipment_group_first_ids (list, optional): Filter by first offline equipment group IDs.
            recipe_ids (list, optional): Filter by specific recipe IDs.
            organism_and_vial_tuples (list, optional): Filter by tuples of organism and vial IDs.
            material_and_lot_tuples (list, optional): Filter by tuples of material and lot IDs.
            gmp (bool, optional): Filter by GMP-compliant processes.
            imported (bool, optional): Filter by imported processes.
            group_ids (list, optional): Filter by specific group IDs.
            master_recipe_ids (list, optional): Filter by master recipe IDs.
            part_of_experiment (bool, optional): Filter by whether processes are part of an experiment.
            sampling_values (list, optional): Filter by specific sampling values.
            sort (list, optional): Sort results (e.g., ['name,asc']).
            page (int, optional): Specify the page number for pagination.
            size (int, optional): Specify the number of results per page (default: 1000).

        Returns:
            List[Dict] or Dict: A list of processes or an error message in JSON format.
        """
        try:
            # Fetch processes using the OpenAPI function with filters applied
            response = get_processes.sync_detailed(
                client=self.client,
                name=name,
                deleted=deleted,
                start_date=start_date,
                end_date=end_date,
                statuses=statuses,
                archived_or_will_be_archived=archived_or_will_be_archived,
                types=types,
                ids=ids,
                lot_ids=lot_ids,
                organism_ids=organism_ids,
                kpi_model_ids=kpi_model_ids,
                material_ids=material_ids,
                stream_ids=stream_ids,
                user_stream_ids=user_stream_ids,
                online_equipment_ids=online_equipment_ids,
                offline_equipment_ids=offline_equipment_ids,
                online_equipment_group_first_ids=online_equipment_group_first_ids,
                offline_equipment_group_first_ids=offline_equipment_group_first_ids,
                recipe_ids=recipe_ids,
                organism_and_vial_tuples=organism_and_vial_tuples,
                material_and_lot_tuples=material_and_lot_tuples,
                gmp=gmp,
                imported=imported,
                group_ids=group_ids,
                master_recipe_ids=master_recipe_ids,
                part_of_experiment=part_of_experiment,
                sampling_values=sampling_values,
                sort=sort,
                page=page,
                size=size,
            )
            print(f"Response Status Code: {response.status_code}")

            # Ensure the response status is OK
            if response.status_code == HTTPStatus.OK:
                return json.loads(response.content.decode())
            else:
                logging.error(
                    f"Failed to fetch processes. HTTP Status: {response.status_code}"
                )
                return {
                    "error": f"Failed to fetch processes. HTTP Status: {response.status_code}"
                }

        except Exception as e:
            logging.error(f"Error fetching processes: {e}")
            return {"error": f"Error fetching processes: {str(e)}"}

    def extract_channels(
        self,
        process_id: int,
        process_phase_id: int = UNSET,
        name: str = UNSET,
        node_types: list = UNSET,
        sensor_type_ids: list = UNSET,
        physical_quantity_unit_id: int = UNSET,
    ):
        """
        Extract channels for a given process ID with optional filters.

        Args:
            process_id (int): ID of the process to fetch channels for.
            process_phase_id (int, optional): Filter by specific process phase ID.
            name (str, optional): Filter by channel name.
            node_types (list, optional): Filter by node types (e.g., ['INPUT', 'OUTPUT']).
            sensor_type_ids (list, optional): Filter by specific sensor type IDs.
            physical_quantity_unit_id (int, optional): Filter by physical quantity unit ID.

        Returns:
            List[Dict] or Dict: A list of channels or an error message in JSON format.
        """
        try:
            # Call the relevant OpenAPI function with the provided arguments
            response = get_channels.sync_detailed(
                process_id=process_id,
                client=self.client,
                process_phase_id=process_phase_id,
                name=name,
                node_types=node_types,
                sensor_type_ids=sensor_type_ids,
                physical_quantity_unit_id=physical_quantity_unit_id,
            )

            # Check if the response status is OK
            if response.status_code == HTTPStatus.OK:
                print(json.dumps(json.loads(response.content.decode()), indent=4))
                return json.loads(response.content.decode())
            else:
                logging.error(
                    f"Failed to fetch channels. HTTP Status: {response.status_code}"
                )
                return {
                    "error": f"Failed to fetch channels. HTTP Status: {response.status_code}"
                }

        except Exception as e:
            logging.error(f"Error fetching channels: {e}")
            return {"error": f"Error fetching channels: {str(e)}"}

    def extract_offline_channels(
        self,
        process_id: int,
    ) -> Dict[str, Any]:
        """
        Extract offline channels for a given process ID.

        Args:
            process_id (int): ID of the process to fetch channels for.

        Returns:
            List[Dict] or Dict: A list of offline channels or an error message in JSON format.
        """
        try:
            # since the endpoint is not part of the public api, call internal api
            base_url = self.client.base_url
            endpoint = f"{base_url}/api/processes/{process_id}/channels/offline"

            offline_channels_response = self.client.get_httpx_client().get(endpoint)

            # Check if the response status is OK
            if offline_channels_response.status_code == HTTPStatus.OK:
                return json.loads(offline_channels_response.content.decode())
            else:
                logging.error(
                    f"Failed to fetch offline channels. HTTP Status: {offline_channels_response.status_code}"
                )
                return {
                    "error": f"Failed to fetch offline channels. HTTP Status: {offline_channels_response.status_code}"
                }

        except Exception as e:
            logging.error(f"Error fetching offline channels: {e}")
            return {"error": f"Error fetching offline channels: {str(e)}"}

    def get_offline_equipment_data(
        self,
        equipment_id: int,
        process_id: int,
    ) -> Dict[str, Any]:
        """
        Get the list of offline data for an offline equipment and a process.

        Args:
            equipment_id (int): ID of the offline equipment.
            process_id (int): ID of the process.

        Returns:
            List[Dict] or Dict: A list of offline data entries with following attributes:
                resolvedSamplingId: the Qubicon or external sample id.
                time: the sample timestamp.
                sensorQubName: The name of the offline sensor.
                value: The value of the offline sensor.
                tag: the name of the sample (for instance: 'day-7')
                discarded: True if the data has been manually discarded
        """
        try:
            # since the endpoint is not part of the public api, call internal api
            base_url = self.client.base_url
            endpoint = f"{base_url}/api/offline-equipment/{equipment_id}/data?processId={process_id}&size=99999"

            offline_data_response = self.client.get_httpx_client().get(endpoint)

            # Check if the response status is OK
            if offline_data_response.status_code == HTTPStatus.OK:
                decoded_response = json.loads(offline_data_response.content.decode())
                return decoded_response["content"]
            else:
                logging.error(
                    f"Failed to fetch offline data. HTTP Status: {offline_data_response.status_code}"
                )
                return {
                    "error": f"Failed to fetch offline data. HTTP Status: {offline_data_response.status_code}"
                }

        except Exception as e:
            logging.error(f"Error fetching offline data: {e}")
            return {"error": f"Error fetching offline data: {str(e)}"}

    def __get_process_start_end_dates(
        self,
        process_id: int,
    ) -> Dict[str, Any]:
        try:
            process_info_resp = self.list_processes(ids=[process_id])
            if (
                not process_info_resp
                or "content" not in process_info_resp
                or not process_info_resp["content"]
            ):
                msg = "No process metadata found to derive start/end dates."
                logging.error(msg)
                return {"error": msg}

            proc_info = process_info_resp["content"][0]
            return proc_info

        except Exception as e:
            logging.error(f"Error getting process start and end dates: {e}")
            return {"error": f"Exception encountered: {str(e)}"}

    def extract_offline_process_data(
        self,
        process_id: int,
        selected_channels: List[Dict[str, Any]],
        start_date: Optional[int] = None,
        end_date: Optional[int] = None,
        granularity: int = 60000,
    ) -> MultiplexChartDataDto:
        """
        Extract offline data from given process and for selected channels.

        Args:
            process_id (int): The ID of the process whose data to extract.
            selected_channels (List[Dict[str, Any]]): A list of dictionaries, each containing 'id' and 'name'.
            start_date (Optional[int], optional): Start time (Unix ms). If None, fetch from process info. Defaults to None.
            end_date (Optional[int], optional): End time (Unix ms). If None, assume now.
            granularity (int, optional): Granularity in ms. Defaults to 60000 (1 minute).

        Returns:
            Dict[str, Any]: A dictionary with either {'success': '...'} or {'error': '...'}.
        """
        try:
            # If start_date and end_date were not provided, fetch from the process details
            if start_date is None or end_date is None:
                proc_info = self.__get_process_start_end_dates(process_id=process_id)
                start_date = proc_info.get("startDate")
                end_date = proc_info.get("endDate")
                if not end_date:
                    end_date = int(time.time() * 1000)

                if not start_date:
                    msg = "Could not determine start date from process metadata."
                    logging.error(msg)
                    return {"error": msg}

            # Construct list of ChannelDataChannelRequestKeyDto objects
            dto_channels = []
            channel_mapping = {}  # Store mapping of ID -> Name for later use
            for ch in selected_channels:
                dto_channels.append(
                    ChannelDataChannelRequestKeyDto(
                        id=ch["id"],
                        type=ChannelDataChannelRequestKeyDtoType.OFFLINE,
                        start_date=start_date,
                        end_date=end_date,
                        granularity=granularity,
                    )
                )
                channel_mapping[ch["id"]] = ch[
                    "name"
                ]  # Save mapping for renaming columns

            # Build the main request DTO
            body = MultiplexChartDataChannelsRequestDto(channels=dto_channels)

            # Call the OpenAPI function
            response = get_multiplex_chart_data_channels.sync_detailed(
                client=self.client, body=body
            )
            if response.status_code != HTTPStatus.OK:
                logging.error(
                    f"Failed to fetch offline process data. HTTP {response.status_code}: {response.content}"
                )
                return {"error": f"Request failed with status {response.status_code}."}

            # Parse response JSON
            data = json.loads(response.content.decode())
            return MultiplexChartDataDto.from_dict(data)

        except Exception as e:
            logging.error(f"Error extracting offline process data: {e}")
            return {"error": f"Exception encountered: {str(e)}"}

    def get_process_samplings(self, process_id: int) -> List[Dict[str, Any]]:
        """
        Fetch detailed sampling data for a given process ID.

        Args:
        process_id: The ID of the process to fetch samplings for

        Returns:
        List of dictionaries containing structured sampling data with:
            - id: Sampling ID
            - value: Sampling value
            - date: Sampling date
            - samplingEventId: Event ID
            - samplingEventType: Type of sampling event
            - lastUsageDate: Last usage timestamp
            - label: Human-readable label
            - discardedValues: List of discarded values
            - externalSamplings: List of external sampling entries
        """
        logging.info(f"Fetching sampling data for process {process_id}")

        try:
            endpoint = f"/api/processes/{process_id}/samplings"
            response = self.client.get_httpx_client().get(
                self.client.base_url.rstrip("/") + endpoint,
                headers={"Authorization": f"Bearer {self.client.token}"},
            )

            if response.status_code != HTTPStatus.OK:
                logging.error(f"Request failed with status {response.status_code}")
                return []

            samplings_raw = response.json().get("content", [])
            structured_samplings = []

            for sampling in samplings_raw:
                structured_sampling = {
                    "id": sampling.get("id"),
                    "value": sampling.get("value"),
                    "date": sampling.get("date"),
                    "samplingEventId": sampling.get("samplingEventId"),
                    "samplingEventType": sampling.get("samplingEventType"),
                    "lastUsageDate": sampling.get("lastUsageDate"),
                    "label": sampling.get("label"),
                    "discardedValues": sampling.get("discardedValues", []),
                    "externalSamplings": [],
                }

                # Process external samplings
                for ext in sampling.get("externalSamplings", []):
                    structured_sampling["externalSamplings"].append(
                        {
                            "id": ext.get("id"),
                            "main": ext.get("main"),
                            "count": ext.get("count"),
                            "value": ext.get("value"),
                            "used": ext.get("used"),
                            "discarded": ext.get("discarded"),
                        }
                    )

                structured_samplings.append(structured_sampling)

            logging.info(f"Retrieved {len(structured_samplings)} samplings")
            return structured_samplings

        except Exception as e:
            logging.error(f"Error fetching process samplings: {str(e)}", exc_info=True)
            return []

    def discard_sampling_data(self, external_sampling_id: str, process_id: int) -> bool:
        """
        Discard a specific external sampling for a given process.

        Args:
        external_sampling_id (str): The external sampling ID to discard. (has to be fetched with get_process_samplings)
        process_id (int): The ID of the process to which the sampling belongs.

        Returns:
        bool: True if the discard was successful, False otherwise.
        """
        logging.info(
            f"Attempting to discard external sampling {external_sampling_id} for process {process_id}"
        )

        endpoint = f"/api/processes/{process_id}/samplings/{external_sampling_id}/externals/{external_sampling_id}/discard"
        url = self.client.base_url.rstrip("/") + endpoint

        payload = {
            "comment": "Discarded via API",
            "reason": "Sample not needed for evaluation",
            "autoStopKpis": False,
        }

        try:
            response = self.client.get_httpx_client().put(
                url,
                headers={"Authorization": f"Bearer {self.client.token}"},
                json=payload,
            )

            if response.status_code == 200:
                logging.info("External sampling discarded successfully.")
                return True
            else:
                logging.error(
                    f"Failed to discard sample. Status code: {response.status_code}"
                )
                logging.error(f"Response: {response.text}")
                return False

        except Exception as e:
            logging.error(
                f"Unexpected error while discarding sampling: {str(e)}", exc_info=True
            )
            return False

    def extract_process_data(
        self,
        process_id: int,
        selected_channels: List[Dict[str, Any]],
        start_date: Optional[int] = None,
        end_date: Optional[int] = None,
        granularity: int = 60000,
        output_file: str = "process_data.json",
        output_format: str = "json",
    ) -> Dict[str, Any]:
        """
        Extract process data for selected channels using the new DTOs and
        save it in JSON or CSV format. Now replaces channel IDs with channel names in CSV.

        If start_date or end_date are not provided, this function will attempt
        to fetch them from the process metadata.

        Args:
            process_id (int): The ID of the process whose data to extract.
            selected_channels (List[Dict[str, Any]]): A list of dictionaries, each containing 'id' and 'name'.
            start_date (Optional[int], optional): Start time (Unix ms). If None, fetch from process info. Defaults to None.
            end_date (Optional[int], optional): End time (Unix ms). If None, fetch from process info. Defaults to None.
            granularity (int, optional): Granularity in ms. Defaults to 60000 (1 minute).
            output_file (str, optional): Path to output file. Defaults to 'process_data.json'.
            output_format (str, optional): 'json' or 'csv'. Defaults to 'json'.

        Returns:
            Dict[str, Any]: A dictionary with either {'success': '...'} or {'error': '...'}.
        """
        try:
            # If start_date and end_date were not provided, fetch from the process details
            if start_date is None or end_date is None:
                process_info_resp = self.list_processes(ids=[process_id])
                if (
                    not process_info_resp
                    or "content" not in process_info_resp
                    or not process_info_resp["content"]
                ):
                    msg = "No process metadata found to derive start/end dates."
                    logging.error(msg)
                    return {"error": msg}

                proc_info = process_info_resp["content"][0]
                if start_date is None:
                    start_date = proc_info.get("startDate")
                if end_date is None:
                    end_date = proc_info.get("endDate")

                if not start_date or not end_date:
                    msg = "Could not determine start/end date from process metadata."
                    logging.error(msg)
                    return {"error": msg}

            # Construct list of ChannelDataChannelRequestKeyDto objects
            dto_channels = []
            channel_mapping = {}  # Store mapping of ID -> Name for later use
            for ch in selected_channels:
                dto_channels.append(
                    ChannelDataChannelRequestKeyDto(
                        id=ch["id"],
                        type=ChannelDataChannelRequestKeyDtoType.ONLINE,
                        start_date=start_date,
                        end_date=end_date,
                        granularity=granularity,
                    )
                )
                channel_mapping[ch["id"]] = ch[
                    "name"
                ]  # Save mapping for renaming columns

            # Build the main request DTO
            body = MultiplexChartDataChannelsRequestDto(channels=dto_channels)

            # Call the OpenAPI function
            response = get_multiplex_chart_data_channels.sync_detailed(
                client=self.client, body=body
            )
            if response.status_code != HTTPStatus.OK:
                logging.error(
                    f"Failed to fetch process data. HTTP {response.status_code}: {response.content}"
                )
                return {"error": f"Request failed with status {response.status_code}."}

            # Parse response JSON
            raw_json = json.loads(response.content.decode())
            channel_array = raw_json.get("channels", [])

            collected_data = []
            for entry in channel_array:
                for val in entry.get("value", []):
                    collected_data.append(
                        {
                            "time": val["time"],
                            "channel_id": entry["key"]["id"],
                            "channel_type": "ONLINE",
                            "value": val["value"],
                        }
                    )

            if not collected_data:
                msg = "No data returned for the given channels."
                logging.warning(msg)
                return {"error": msg}

            # Convert results to DataFrame
            df = pd.DataFrame(collected_data)

            # Decide output format
            output_format = output_format.lower()
            if output_format == "json":
                with open(output_file, "w") as f:
                    json.dump(collected_data, f, indent=4)
                logging.info(
                    f"Process data successfully written to {output_file} as JSON"
                )
                return {"success": f"Data exported to {output_file} as JSON"}
            elif output_format == "csv":
                # Pivot: time as index, channel_id as columns, values as cells
                pivot_df = df.pivot(index="time", columns="channel_id", values="value")

                # Rename channel ID columns with their actual names
                pivot_df.rename(columns=channel_mapping, inplace=True)

                pivot_df.to_csv(output_file)
                logging.info(
                    f"Process data successfully written to {output_file} as CSV with channel names"
                )
                return {"success": f"Data exported to {output_file} as CSV"}
            else:
                msg = f"Invalid output format: {output_format}"
                logging.error(msg)
                return {"error": msg}

        except Exception as e:
            logging.error(f"Error extracting process data: {e}")
            return {"error": f"Exception encountered: {str(e)}"}

    def get_process_tag_values(self, process_id: int) -> List[Dict[str, Any]]:
        """
        Fetches process tag values and extracts TagDto with corresponding values.

        Args:
            process_id: The process ID to fetch tag values for

        Returns:
            List of dictionaries containing TagDto information and values
            [{
                'tag_value_id': int,  # The correct ID for updates (top-level id)
                'tag_id': int,        # The tag definition ID (nested tag.id)
                'tag_name': str,
                'description': str,
                'status': str,
                'scope': str,
                'physical_quantity_unit': dict,
                'value': any,
                'adhoc': bool,
                'initial_value': any,
                'update_date': str,
                'calculate_date': str,
                'raw_data': dict      # Complete raw response data
            }]
        """
        endpoint = f"/api/processes/{process_id}/tag-values"

        try:
            response = self.client.get_httpx_client().get(
                self.client.base_url + endpoint,
                headers={"Authorization": f"Bearer {self.client.token}"},
            )

            if response.status_code != HTTPStatus.OK:
                logging.error(f"Request failed with status {response.status_code}")
                return []

            data = response.json()
            tag_values = []

            for item in data.get("content", []):
                tag = item.get("tag", {})
                tag_values.append(
                    {
                        "tag_value_id": item.get("id"),
                        "tag_id": tag.get("id"),
                        "tag_name": tag.get("name"),
                        "description": tag.get("description"),
                        "status": tag.get("status"),
                        "scope": tag.get("scope"),
                        "physical_quantity_unit": tag.get("physicalQuantityUnit", {}),
                        "value": item.get("value"),
                        "adhoc": item.get("adhoc"),
                        "initial_value": item.get("initialValue"),
                        "update_date": item.get("updateDate"),
                        "calculate_date": item.get("calculateDate"),
                        "raw_data": item,
                    }
                )

            return tag_values

        except Exception as e:
            logging.error(f"Error fetching tag values: {str(e)}")
            return []

    def set_tag_value(
        self, process_id: int, tag_value_id: int, new_value: Any
    ) -> Optional[Dict[str, Any]]:
        """
        Updates a tag value for a specific process and tag value ID.

        Args:
            process_id: The process ID
            tag_value_id: The tag value ID to update (top-level id)
            new_value: The new value to set

        Returns:
            The updated ProcessTagValueDto if successful, None otherwise
        """
        endpoint = f"/api/processes/{process_id}/tag-values/{tag_value_id}/update-value"

        try:
            response = self.client.get_httpx_client().put(
                self.client.base_url + endpoint,
                headers={
                    "Authorization": f"Bearer {self.client.token}",
                    "Content-Type": "application/json",
                },
                json={"value": new_value},
            )

            if response.status_code == HTTPStatus.OK:
                logging.info(f"Successfully updated tag value {tag_value_id}")
                return response.json()
            else:
                logging.error(
                    f"Failed to update tag value. Status: {response.status_code}"
                )
                logging.debug(f"Response: {response.text}")
                return None

        except Exception as e:
            logging.error(f"Error updating tag value: {str(e)}")
            return None

    def stream_events(self, event_types: List[str], nonce: str):
        """Stream events from the API."""
        try:
            response = stream_events.sync(
                client=self.client, types=event_types, nonce=nonce
            )
            if response:
                for event in response:
                    print(f"Event: {event}")
            else:
                print("No events streamed.")
        except Exception as e:
            print(f"Error streaming events: {e}")

    def convert_model_to_importable_json(self, api_response):
        """
        Converts the API response of a model into an importable JSON structure.

        Args:
            api_response (dict): The API response from the GET model details endpoint.

        Returns:
            dict: The importable JSON structure.
        """
        if not api_response:
            logging.error("No valid model data to convert.")
            return None

        return {
            "engineType": api_response.get("engineType"),
            "kpiName": api_response.get("kpiName"),
            "abbr": api_response.get("abbr"),
            "calculationStyle": api_response.get("calculationStyle"),
            "status": "DRAFT",
            "description": api_response.get("description", ""),
            "script": api_response.get("script", ""),
            "inputs": [
                {
                    "name": input_item.get("name"),
                    "order": idx,
                    "physicalQuantityUnit": input_item.get("physicalQuantityUnit", {}),
                    "description": input_item.get("description", ""),
                }
                for idx, input_item in enumerate(api_response.get("inputs", []))
            ],
            "outputs": [
                {
                    "name": output_item.get("name"),
                    "order": idx,
                    "physicalQuantityUnit": output_item.get("physicalQuantityUnit", {}),
                    "description": output_item.get("description", ""),
                }
                for idx, output_item in enumerate(api_response.get("outputs", []))
            ],
        }

    def export_model_to_json(self, model_id: int, filename: str):
        """Export model details to a JSON file."""
        model_details = self.fetch_model_details(model_id)

        if not model_details:
            logging.error("Model details not found or could not be exported.")
            return

        importable_json = self.convert_model_to_importable_json(model_details)

        if not importable_json:
            logging.error("Failed to convert model to importable format.")
            return

        try:
            with open(filename, "w") as file:
                json.dump(model_details, file, indent=4)
            logging.info(f"Model details exported to '{filename}' successfully.")
        except Exception as e:
            logging.error(f"Error exporting model details: {e}")

    def create_computable_model(self, model_data: dict) -> Optional[Dict[str, Any]]:
        """
        Create a new computable model using the provided JSON data,
        ensuring we handle all Enums and sub-DTOs manually so that
        there's no 'str' object has no attribute 'value' error.
        """
        try:
            # ------------------------------------------------------
            # Step 1: Basic cleanup & mandatory checks
            # ------------------------------------------------------
            if "engineType" not in model_data:
                raise ValueError("Missing 'engineType' in model data.")

            # Remove fields that must NOT be submitted
            for ro_field in ["id", "creationDate", "updateDate"]:
                model_data.pop(ro_field, None)

            # If sensorType is None, remove it
            if model_data.get("sensorType") is None:
                model_data.pop("sensorType", None)

            # ------------------------------------------------------
            # Step 2: Handle physical quantities (existing logic)
            # ------------------------------------------------------
            model_data = self.handle_physical_quantities(model_data)
            if not model_data:
                raise ValueError("Failed to process physical quantities.")

            # ------------------------------------------------------
            # Step 3: Manually build all sub-dtos (inputs/outputs)
            # ------------------------------------------------------
            input_dto_list = []
            for inp in model_data.get("inputs", []):
                # Safely build PhysicalQuantityUnitDto if present
                pq_unit_dict = inp.get("physicalQuantityUnit", {})
                # Convert "status" if it exists
                raw_pq_status = pq_unit_dict.get("status", UNSET)
                if raw_pq_status not in [UNSET, None]:
                    raw_pq_status = PhysicalQuantityUnitDtoStatus(raw_pq_status)

                pq_unit_dto = PhysicalQuantityUnitDto(
                    unit=pq_unit_dict.get("unit", ""),
                    name=pq_unit_dict.get("name", ""),
                    id=pq_unit_dict.get("id", UNSET),
                    physical_quantity_id=pq_unit_dict.get("physicalQuantityId", UNSET),
                    status=raw_pq_status,
                )

                # Build the ComputableModelInputDto (note: no publishField here)
                input_dto = ComputableModelInputDto(
                    name=inp.get("name", ""),
                    physical_quantity_unit=pq_unit_dto,
                    id=inp.get("id", UNSET),
                    description=inp.get("description", UNSET),
                    order=inp.get("order", UNSET),
                )
                input_dto_list.append(input_dto)

            output_dto_list = []
            for outp in model_data.get("outputs", []):
                # If "publishField" is None, remove it so we don't break the constructor
                if outp.get("publishField") is None:
                    outp.pop("publishField", None)

                pq_unit_dict = outp.get("physicalQuantityUnit", {})
                # Convert "status" if it exists
                raw_pq_status = pq_unit_dict.get("status", UNSET)
                if raw_pq_status not in [UNSET, None]:
                    raw_pq_status = PhysicalQuantityUnitDtoStatus(raw_pq_status)

                pq_unit_dto = PhysicalQuantityUnitDto(
                    unit=pq_unit_dict.get("unit", ""),
                    name=pq_unit_dict.get("name", ""),
                    id=pq_unit_dict.get("id", UNSET),
                    physical_quantity_id=pq_unit_dict.get("physicalQuantityId", UNSET),
                    status=raw_pq_status,
                )

                # Build the ComputableModelOutputDto
                output_dto = ComputableModelOutputDto(
                    name=outp.get("name", ""),
                    physical_quantity_unit=pq_unit_dto,
                    id=outp.get("id", UNSET),
                    description=outp.get("description", UNSET),
                    order=outp.get("order", UNSET),
                    # If you want a publishField, pass it here; if not, do UNSET
                    publish_field=UNSET,
                )
                output_dto_list.append(output_dto)

            # ------------------------------------------------------
            # Step 4: Convert top-level strings into Enums, build final DTO
            # ------------------------------------------------------
            # engineType => AbstractComputableModelDtoEngineType
            engine_type_str = model_data[
                "engineType"
            ]  # e.g. "JYTHON" or "EXTERNAL_PYTHON"
            engine_type_enum = AbstractComputableModelDtoEngineType(engine_type_str)

            # calculationStyle => AbstractComputableModelDtoCalculationStyle
            calc_style_str = model_data.get("calculationStyle", "ONLINE")
            calc_style_enum = AbstractComputableModelDtoCalculationStyle(calc_style_str)

            # status => AbstractComputableModelDtoStatus
            status_str = model_data.get("status", "DRAFT")
            status_enum = AbstractComputableModelDtoStatus(status_str)

            # Choose which top-level class to build
            if engine_type_enum == AbstractComputableModelDtoEngineType.JYTHON:
                model_body = JythonComputableModelDto(
                    engine_type=engine_type_enum,
                    kpi_name=model_data.get("kpiName", ""),
                    abbr=model_data.get("abbr", ""),
                    calculation_style=calc_style_enum,
                    outputs=output_dto_list,
                    inputs=input_dto_list,  # pass your built inputs
                    status=status_enum,
                    description=model_data.get("description", UNSET),
                    script=model_data.get("script", UNSET),
                    # creation_date / update_date can remain UNSET
                )
            else:
                # EXTERNAL_PYTHON or other
                model_body = ExternalPythonComputableModelDto(
                    engine_type=engine_type_enum,
                    kpi_name=model_data.get("kpiName", ""),
                    abbr=model_data.get("abbr", ""),
                    calculation_style=calc_style_enum,
                    outputs=output_dto_list,
                    inputs=input_dto_list,
                    status=status_enum,
                    description=model_data.get("description", UNSET),
                    script=model_data.get("script", UNSET),
                )

            # ------------------------------------------------------
            # Step 5: Call the API
            # ------------------------------------------------------
            response = create_computable_model.sync_detailed(
                client=self.client,
                body=model_body,
            )

            if response.status_code == HTTPStatus.OK:
                parsed = json.loads(response.content.decode())
                logging.info("Successfully created computable model.")
                return parsed
            elif response.status_code == HTTPStatus.BAD_REQUEST:
                # If the server returns 400 with valid.unique.kpi.name, log a friendlier message
                if "valid.unique.kpi.name" in json.loads(response.content.decode()):
                    logging.error(
                        "Failed to create model because 'kpiName' is already in use. "
                        "Please choose a different kpiName."
                    )
            else:
                msg = (
                    f"Failed to create computable model. "
                    f"HTTP Status: {response.status_code}. "
                    f"Response: {response.content.decode()}"
                )
                logging.error(msg)
                return None

        except Exception as e:
            logging.error(f"Error creating computable model: {e}")
            return None

    def create_physical_quantity(
        self, pq_dict: Dict[str, Any]
    ) -> Optional[Dict[str, Any]]:
        """
        Create a new physical quantity (or possibly retrieve it if the server
        returns 200 instead of 201). We rely on the server's logic to handle
        duplicates by name.

        Returns the parsed JSON on success, or None on error.
        """
        try:
            # Convert input dict to the generated DTO.
            # Make sure the keys in pq_dict match PhysicalQuantityDto’s fields (name, units, etc.).
            dto = PhysicalQuantityDto.from_dict(pq_dict)

            # Call the generated POST function
            response = create_physical_quantity.sync_detailed(
                client=self.client,
                body=dto,
            )

            # Some servers return 201 for “created”, some might return 200 if it already existed.
            if response.status_code in (HTTPStatus.CREATED, HTTPStatus.OK):
                # Manually parse the body into a dict
                return json.loads(response.content.decode())
            else:
                logging.error(
                    f"Failed to create physical quantity. "
                    f"HTTP Status: {response.status_code}. "
                    f"Response: {response.content.decode()}"
                )
                return None

        except Exception as e:
            logging.error(f"Error creating physical quantity: {e}")
            return None

    def check_existing_physical_quantity(self, pq_name: str) -> Optional[dict]:
        """
        Check if a physical quantity already exists in the system.

        Args:
            pq_name (str): Name of the physical quantity to check.

        Returns:
            dict or None: The physical quantity data if found, or None if not found.
        """
        try:
            response = get_physical_quantities.sync_detailed(client=self.client)
            if response.status_code in (HTTPStatus.OK, HTTPStatus.CREATED):
                parsed_response = json.loads(response.content.decode())
                for pq in parsed_response:
                    if self.fuzzy_match(pq["name"], pq_name):
                        return pq
                return None
            else:
                logging.error(
                    f"Failed to fetch physical quantities. HTTP Status: {response.status_code}"
                )
                return None
        except Exception as e:
            logging.error(f"Error fetching physical quantities: {e}")
            return None

    def handle_physical_quantities(self, model_data: dict) -> Optional[dict]:
        """
        Handle physical quantities and map them correctly for both inputs and outputs.

        Args:
            model_data (dict): Model data containing inputs and outputs with physical quantities.

        Returns:
            dict or None: Updated model data or None if mapping fails.
        """
        try:
            for var in model_data["inputs"] + model_data["outputs"]:
                pq_name = var["physicalQuantityUnit"]["name"]
                unit_name = var["physicalQuantityUnit"]["unit"]

                # Check if the physical quantity already exists
                existing_pq = self.check_existing_physical_quantity(pq_name)

                if existing_pq:
                    logging.info(
                        f"Found existing physical quantity '{pq_name}' (ID: {existing_pq['id']})."
                    )
                    var["physicalQuantityUnit"]["physicalQuantityId"] = existing_pq[
                        "id"
                    ]

                    # Check if the unit exists within the physical quantity
                    existing_unit = self.check_existing_unit(existing_pq, unit_name)

                    if existing_unit:
                        var["physicalQuantityUnit"]["unit"] = existing_unit["unit"]
                        var["physicalQuantityUnit"]["id"] = existing_unit["id"]
                    else:
                        updated_pq = self.add_unit_to_existing_physical_quantity(
                            existing_pq["id"], pq_name, unit_name
                        )
                        if updated_pq:
                            new_unit = next(
                                (
                                    u
                                    for u in updated_pq["units"]
                                    if u["unit"] == unit_name
                                ),
                                None,
                            )
                            if new_unit:
                                var["physicalQuantityUnit"]["unit"] = new_unit["unit"]
                                var["physicalQuantityUnit"]["id"] = new_unit["id"]
                            else:
                                raise ValueError(
                                    f"Failed to retrieve the new unit '{unit_name}'."
                                )
                else:
                    # Create a new physical quantity
                    new_pq = self.create_physical_quantity(
                        {"name": pq_name, "units": [{"unit": unit_name}]}
                    )
                    if new_pq:
                        var["physicalQuantityUnit"]["physicalQuantityId"] = new_pq["id"]
                        new_unit = new_pq["units"][0]
                        var["physicalQuantityUnit"]["unit"] = new_unit["unit"]
                        var["physicalQuantityUnit"]["id"] = new_unit["id"]
                    else:
                        raise ValueError(
                            f"Failed to create physical quantity '{pq_name}'."
                        )

            return model_data
        except Exception as e:
            logging.error(f"Error handling physical quantities: {e}")
            return None

    def check_existing_unit(
        self, physical_quantity: dict, unit_name: str
    ) -> Optional[dict]:
        """
        Match a unit within a physical quantity.

        Args:
            physical_quantity (dict): The physical quantity data.
            unit_name (str): The name of the unit to match.

        Returns:
            dict or None: Matching unit or None if not found.
        """
        for unit in physical_quantity["units"]:
            if self.fuzzy_match(unit["unit"], unit_name):
                logging.info(
                    f"Unit match found: '{unit_name}' within physical quantity '{physical_quantity['name']}' (ID: {unit['id']})."
                )
                return unit
        logging.info(
            f"No unit match found for '{unit_name}' in physical quantity '{physical_quantity['name']}'."
        )
        return None

    def normalize_name(self, name: str) -> str:
        """Normalize physical quantity or unit names."""
        return re.sub(r"[^a-zA-Z0-9]", "", name).lower()

    def fuzzy_match(self, name1: str, name2: str, threshold: int = 93) -> bool:
        """Perform fuzzy matching between two names."""
        normalized_name1 = self.normalize_name(name1)
        normalized_name2 = self.normalize_name(name2)
        similarity = fuzz.token_sort_ratio(normalized_name1, normalized_name2)
        return similarity >= threshold

    def add_unit_to_existing_physical_quantity(
        self, pq_id: int, pq_name: str, unit_name: str
    ) -> Optional[Dict[str, Any]]:
        """
        Work around the absence of an 'update' method by:
          1) Fetching the existing physical quantity (via check_existing_physical_quantity).
          2) Appending the new unit to the 'units' array.
          3) Calling create_physical_quantity(...) again with the updated data.
             This assumes the backend merges or updates based on PQ 'name'.

        Returns the updated PQ data on success or None on failure.
        """
        # 1) Get the existing physical quantity by name
        existing_pq = self.check_existing_physical_quantity(pq_name)
        if not existing_pq:
            logging.error(f"Physical quantity '{pq_name}' not found for updating.")
            return None

        logging.info(
            f"Adding unit '{unit_name}' to existing physical quantity '{pq_name}' (ID: {pq_id})."
        )

        # 2) Append new unit to the array
        #    existing_pq['units'] is a list of dicts like [{"unit": "L", "physicalQuantityId": X}, ...]
        units = existing_pq.get("units", [])
        units.append({"unit": unit_name, "physicalQuantityId": pq_id})

        # 3) Construct data that we will pass back into create_physical_quantity
        #    We typically keep name, status, and all units.
        updated_data = {
            "name": pq_name,
            "status": existing_pq.get("status", "IN_USE"),
            "units": units,
        }
        # NOTE: The "id" field might or might not break the re-POST—depends on backend rules.
        # Sometimes it's safest to omit "id" because the server may reject duplicates.
        # But if your backend accepts an "id" on create, you can include it:
        #
        # updated_data["id"] = pq_id

        # Now call create_physical_quantity(...) again
        updated_pq = self.create_physical_quantity(updated_data)
        if updated_pq:
            logging.info(f"Unit '{unit_name}' added successfully to '{pq_name}'.")
            return updated_pq
        else:
            logging.error(
                f"Failed to add unit '{unit_name}' to physical quantity '{pq_name}'."
            )
            return None

    def list_recipes(
        self,
        page: int = 0,
        size: int = 400,
        type: str = "NORMAL",
        statuses: List[str] = ["DRAFT", "RELEASED", "GMP"],
        sort: List[str] = ["updateDate,desc"],
        search: str = "",
    ) -> Optional[Dict[str, Any]]:
        """
        List all recipes with optional query parameters.

        Args:
            page (int, optional): Page number. Default=0
            size (int, optional): Page size. Default=400
            type (str, optional): Filter by recipe type. Default='NORMAL'
            statuses (List[str], optional): Filter by statuses. Default=['DRAFT','RELEASED','GMP']
            sort (List[str], optional): Sorting fields. Default=['updateDate,desc']
            search (str, optional): Search string. Default=''

        Returns:
            Dict or None: A dict containing recipes or None if request fails.
        """
        base_url = self.client.base_url  # Or however you retrieve the base URL
        endpoint = f"{base_url}/api/control-recipes"
        params = {
            "page": page,
            "size": size,
            "type": type,
            "statuses": ",".join(statuses),
            "sort": sort,
            "search": search,
        }

        try:
            response = self.client.get_httpx_client().get(endpoint, params=params)
            if response.status_code == HTTPStatus.OK:
                return response.json()
            else:
                logging.error(
                    f"Failed to list recipes. HTTP {response.status_code}: {response.text}"
                )
                return None
        except Exception as e:
            logging.error(f"Error listing recipes: {e}")
            return None

    def get_recipe(self, recipe_id: int) -> Optional[Dict[str, Any]]:
        """
        Fetch a single recipe by its ID.

        Args:
            recipe_id (int): The unique ID of the recipe.

        Returns:
            Dict or None: A dict with recipe details, or None on failure.
        """
        base_url = self.client.base_url
        endpoint = f"{base_url}/api/control-recipes/{recipe_id}"
        params = {"deleted": "false"}  # from your example

        try:
            response = self.client.get_httpx_client().get(endpoint, params=params)
            if response.status_code == HTTPStatus.OK:
                return response.json()
            else:
                logging.error(
                    f"Failed to get recipe {recipe_id}. HTTP {response.status_code}: {response.text}"
                )
                return None
        except Exception as e:
            logging.error(f"Error getting recipe {recipe_id}: {e}")
            return None

    def create_process(
        self,
        name: str,
        recipe_id: int,
        description: str = "",
        process_type: str = "NORMAL",
        online_equipment_ids: Optional[List[int]] = None,
        offline_equipment_ids: Optional[List[int]] = None,
        process_materials: Optional[List[Any]] = None,
        process_organisms: Optional[List[Any]] = None,
        streams: Optional[List[Any]] = None,
        alert_users: Optional[List[Any]] = None,
    ) -> Optional[Dict[str, Any]]:
        """
        Create a new process by POSTing to /api/processes.

        Args:
            name (str): The name of the new process.
            recipe_id (int): The recipe ID on which to base this process (aka processDefinitionId).
            description (str, optional): A short description of the process.
            process_type (str, optional): Type of process (usually 'NORMAL').
            online_equipment_ids (List[int], optional): Any associated online equipment IDs.
            offline_equipment_ids (List[int], optional): Any associated offline equipment IDs.
            process_materials (List, optional): If you want to pass any material references.
            process_organisms (List, optional): If you want to pass organism references.
            streams (List, optional): Any streams for the process.
            alert_users (List, optional): IDs or references for users to alert.

        Returns:
            Dict or None: The newly created process as a dict, or None on error.
        """
        base_url = self.client.base_url
        endpoint = f"{base_url}/api/processes"

        payload = {
            "name": name,
            "description": description,
            "type": process_type,
            "processDefinitionId": recipe_id,  # i.e. 'recipe' id
            "onlineEquipmentIds": online_equipment_ids or [],
            "offlineEquipmentIds": offline_equipment_ids or [],
            "processMaterials": process_materials or [],
            "processOrganisms": process_organisms or [],
            "streams": streams or [],
            "alertUsers": alert_users or [],
        }

        try:
            response = self.client.get_httpx_client().post(endpoint, json=payload)
            if 200 <= response.status_code < 300:
                return response.json()  # Return the newly created process
            else:
                logging.error(
                    f"Failed to create process. HTTP {response.status_code}: {response.text}"
                )
                return None
        except Exception as e:
            logging.error(f"Error creating process: {e}")
            return None

    def start_process(
        self,
        process_id: int,
    ) -> bool:
        """
        Start a process.

        Args:
            process_id (int): The id of the process to start.

        Returns:
            True or False on error.
        """
        base_url = self.client.base_url
        endpoint = f"{base_url}/api/processes/{process_id}/start"

        try:
            response = self.client.get_httpx_client().post(endpoint)
            if 200 <= response.status_code < 300:
                return True
            else:
                logging.error(
                    f"Failed to start process. HTTP {response.status_code}: {response.text}"
                )
                return False
        except Exception as e:
            logging.error(f"Error starting process: {e}")
            return False

    def create_process_group(self, name: str) -> Optional[Dict[str, Any]]:
        """
        Create a new process group by POSTing to /api/process-groups.

        Args:
            name (str): Name of the process group.
            status (str, optional): Group status. Typically 'RELEASED'.

        Returns:
            Dict or None: The newly created group or None on error.
        """
        base_url = self.client.base_url
        endpoint = f"{base_url}/api/process-groups"

        body = {"name": name, "status": "RELEASED"}

        try:
            response = self.client.get_httpx_client().post(endpoint, json=body)
            if 200 <= response.status_code < 300:
                return response.json()
            else:
                logging.error(
                    f"Failed to create process group. HTTP {response.status_code}: {response.text}"
                )
                return None
        except Exception as e:
            logging.error(f"Error creating process group: {e}")
            return None

    def add_processes_to_group(
        self, group_id: int, process_ids: List[int]
    ) -> Optional[Dict[str, Any]]:
        """
        Add one or more processes to an existing group by PUTting the entire group object
        including the updated 'processes' array.

        Args:
            group_id (int): The existing process group ID.
            process_ids (List[int]): A list of process IDs to include in this group.

        Returns:
            Dict or None: Updated group data or None on error.
        """
        base_url = self.client.base_url
        endpoint = f"{base_url}/api/process-groups/{group_id}"

        # 1) Fetch the current group object
        try:
            current_group_resp = self.client.get_httpx_client().get(endpoint)
            if current_group_resp.status_code != HTTPStatus.OK:
                logging.error(
                    f"Failed to fetch group {group_id}. HTTP {current_group_resp.status_code}: {current_group_resp.text}"
                )
                return None
            group_data = current_group_resp.json()
        except Exception as e:
            logging.error(f"Error fetching group {group_id} before update: {e}")
            return None

        # 2) Update the 'processes' list with the new IDs
        existing_processes = group_data.get("processes", [])
        # each item typically looks like {"id": 2876, "name": "...", "status": "READY", ...}

        # Convert current IDs to a set for quick membership checks
        existing_ids = set(p["id"] for p in existing_processes if "id" in p)

        for pid in process_ids:
            if pid not in existing_ids:
                # Append minimal structure that the backend expects
                existing_processes.append({"id": pid, "status": "READY"})

        group_data["processes"] = existing_processes

        # 3) PUT the entire group object back
        try:
            put_response = self.client.get_httpx_client().put(endpoint, json=group_data)
            if 200 <= put_response.status_code < 300:
                return put_response.json()
            else:
                logging.error(
                    f"Failed to add processes to group {group_id}. "
                    f"HTTP {put_response.status_code}: {put_response.text}"
                )
                return None
        except Exception as e:
            logging.error(f"Error updating group {group_id}: {e}")
            return None

    def delete_process(self, process_id: int) -> Optional[Dict[str, Any]]:
        """
        Delete a process by its ID.

        Args:
            process_id (int): The ID of the process to delete.

        Returns:
            Dict or None: The response JSON (if available) or None.
        """
        base_url = self.client._base_url  # Ensure this is correctly set
        endpoint = f"{base_url}/api/processes/{process_id}"

        try:
            response = self.client.get_httpx_client().delete(endpoint)

            if response.status_code in {200, 204}:  # Handle both success cases
                if response.status_code == 204 or not response.content.strip():
                    logging.info(
                        f"Process {process_id} deleted successfully. No content returned."
                    )
                    return {
                        "message": f"Process {process_id} deleted successfully."
                    }  # Simulated response
                return response.json()  # If JSON is available, return it

            logging.error(
                f"Failed to delete process {process_id}. HTTP {response.status_code}: {response.text}"
            )
            return None

        except Exception as e:
            logging.error(f"Error deleting process {process_id}: {e}")
            return None

    def delete_multiple_processes(
        self, process_ids: List[int]
    ) -> Dict[int, Optional[Dict[str, Any]]]:
        """
        Delete multiple processes by their IDs.

        Args:
            process_ids (List[int]): A list of process IDs to delete.

        Returns:
            Dict[int, Optional[Dict[str, Any]]]: A dictionary mapping process IDs to their deletion response or None if failed.
        """
        results = {}
        for process_id in process_ids:
            results[process_id] = self.delete_process(
                process_id
            )  # Call existing function
        return results

    def start_processgroup(self, group_id: int) -> bool:
        """
        Start a process group.

        Args:
            group_id (int): The ID of the process group to start.

        Returns:
            bool: True if successful, False otherwise.
        """
        base_url = self.client.base_url
        http_client = self.client.get_httpx_client()

        # Step 1: Fetch group details
        group_endpoint = f"{base_url}/api/process-groups/{group_id}"
        try:
            group_resp = http_client.get(group_endpoint)
            if group_resp.status_code != HTTPStatus.OK:
                logging.error(
                    f"Failed to fetch group {group_id}. HTTP {group_resp.status_code}: {group_resp.text}"
                )
                return False

            group_data = group_resp.json()
            logging.info(f"Fetched group data for group {group_id}: {group_data}")
        except Exception as e:
            logging.error(f"Error fetching group {group_id}: {e}")
            return False

        # Step 2: Extract and filter process IDs
        existing_processes = group_data.get("processes", [])
        process_ids = [
            p["id"] for p in existing_processes if p.get("status") == "READY"
        ]

        if not process_ids:
            logging.error(
                f"No processes in READY state for group {group_id}. Cannot start."
            )
            return False

        json_payload = {"processIds": process_ids}

        # Step 3: Call the endpoint
        endpoint = f"{base_url}/api/process-groups/{group_id}/start"

        try:
            response = http_client.post(endpoint, json=json_payload)
            if not 200 <= response.status_code < 300:
                logging.error(
                    f"Failed at Group start: {response.status_code} - {response.text}"
                )
                return False

            logging.info(f"Successfully started process group {group_id}")
            return True
        except Exception as e:
            logging.error(f"Error during call to {endpoint}: {e}")
            return False

    def stop_processgroup(self, group_id: int) -> bool:
        """
        Stop a process group.

        Args:
            group_id (int): The ID of the process group to stop.

        Returns:
            bool: True if successful, False otherwise.
        """
        base_url = self.client.base_url
        http_client = self.client.get_httpx_client()

        # Call the endpoint
        endpoint = f"{base_url}/api/process-groups/{group_id}/stop"

        try:
            response = http_client.post(endpoint)
            if not 200 <= response.status_code < 300:
                logging.error(
                    f"Failed at Group stop: {response.status_code} - {response.text}"
                )
                return False

            logging.info(f"Successfully stopped process group {group_id}")
            return True
        except Exception as e:
            logging.error(f"Error during call to {endpoint}: {e}")
            return False

    def get_process_state(self, process_id: int) -> Dict[str, Any]:
        """
        Fetch the current state for a specific process ID.

        Args:
            process_id: The ID of the process to fetch state for

        Returns:
            Dictionary containing the process state information. The exact structure
            depends on the API response, but typically includes:
            - status: Current status of the process
            - progress: Completion progress (if applicable)
            - timestamps: Various process timestamps
            - other process-specific state information

        Raises:
            ValueError: If process_id is not provided or invalid
        """
        if not isinstance(process_id, int) or process_id < 0:
            raise ValueError("process_id must be a positive integer")

        logging.info(f"Fetching state for process {process_id}")

        try:
            endpoint = f"/api/processes/{process_id}/state"
            response = self.client.get_httpx_client().get(
                self.client.base_url.rstrip("/") + endpoint,
                headers={"Authorization": f"Bearer {self.client.token}"},
            )

            if response.status_code == HTTPStatus.NOT_FOUND:
                logging.warning(f"Process {process_id} not found")
                return {}
            elif response.status_code != HTTPStatus.OK:
                logging.error(f"Request failed with status {response.status_code}")
                return {}

            state_data = response.json()
            logging.info(f"Process {process_id} state data: {state_data}")

            return state_data

        except Exception as e:
            logging.error(
                f"Error fetching state for process {process_id}: {str(e)}",
                exc_info=True,
            )
            return {}

    def get_active_processes_list_state(
        self,
        statuses: List[str] = ["READY", "RUNNING", "WARMING_UP"],
    ) -> List[Dict[str, Any]]:
        """
        Fetch list of active processes.

        Args:
            statuses: List of active statuses to filter by (default: ["READY", "RUNNING", "WARMING_UP"])

        Returns:
            List of dictionaries containing structured process status data with all fields
            from the API response.
        """
        logging.info("Fetching filtered process status data")

        try:
            endpoint = "/api/processes/partial/minimum"
            params = {
                "archivedOrWillBeArchived": "false",
                "sort": "id,desc",
                "page": "0",
                "size": 999,
                "statuses": ",".join(statuses),
            }

            response = self.client.get_httpx_client().get(
                self.client.base_url.rstrip("/") + endpoint,
                headers={"Authorization": f"Bearer {self.client.token}"},
                params=params,
            )

            if response.status_code != HTTPStatus.OK:
                logging.error(f"Request failed with status {response.status_code}")
                return []

            response_data = response.json()

            # Log all information from the response
            logging.info(f"Full response data with filters: {response_data}")

            # Handle paginated response (common with page/size parameters)
            if isinstance(response_data, dict):
                if "content" in response_data:
                    # Assuming paginated response with content field
                    return response_data.get("content", [])
                else:
                    # If it's a dictionary but not paginated, wrap in list
                    return [response_data]
            elif isinstance(response_data, list):
                return response_data
            else:
                logging.warning(f"Unexpected response format: {type(response_data)}")
                return []

        except Exception as e:
            logging.error(
                f"Error fetching filtered process status: {str(e)}", exc_info=True
            )
            return []

    def generate_and_upload_sampling_plan(
        self, process_group_id: int, number_of_samples: int, sample_prefix: str = "Day"
    ) -> bool:
        """
        Generates and uploads a sampling plan to the specified process group ID.

        Args:
            process_group_id: The ID of the process group to upload to
            number_of_samples: Number of samples to generate for each process
            sample_prefix: Prefix for the sampling schedule (e.g., "Day")

        Returns:
            bool: True if upload was successful, False otherwise
        """
        # Validate input
        if not isinstance(process_group_id, int) or process_group_id <= 0:
            logging.error(f"Invalid process group ID provided: {process_group_id}")
            return False

        logging.info(f"Starting sampling plan generation for group {process_group_id}")

        # 1. Fetch the process group details
        try:
            endpoint = f"{self.client.base_url.rstrip('/')}/api/process-groups/{process_group_id}"
            logging.info(f"Fetching process group details from: {endpoint}")
            response = self.client.get_httpx_client().get(endpoint)

            if response.status_code != 200:
                logging.error(
                    f"Failed to fetch group details. Status: {response.status_code}"
                )
                logging.error(f"Response: {response.text}")
                return False

            process_group = response.json()
            if "processes" not in process_group:
                logging.error("Process group data doesn't contain 'processes' field")
                return False

        except Exception as e:
            logging.error(f"Error fetching process group details: {str(e)}")
            return False

        # 2. Generate Excel file
        try:
            logging.info("Generating Excel file")
            wb = Workbook()
            ws = wb.active

            # Set headers
            headers = ["Process", "Sampling schedule", "External ID"]
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                ws[f"{col_letter}1"] = header
                ws[f"{col_letter}1"].font = Font(bold=True)

            # Populate data
            current_row = 2
            process_count = len(process_group["processes"])
            sample_count = process_count * number_of_samples
            logging.info(
                f"Adding {process_count} processes with {number_of_samples} samples each (total: {sample_count})"
            )

            for process in process_group["processes"]:
                name = process.get("name", "Unnamed Process")
                for i in range(1, number_of_samples + 1):
                    ws[f"A{current_row}"] = name
                    ws[f"B{current_row}"] = f"{sample_prefix} {i}"
                    ws[f"C{current_row}"] = f"{name}_{i:03d}"
                    current_row += 1

            # Auto-adjust columns
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column].width = (max_length + 2) * 1.2

            # Save to memory
            excel_stream = BytesIO()
            wb.save(excel_stream)
            excel_stream.seek(0)
            file_size = len(excel_stream.getvalue()) / 1024  # KB
            logging.info(f"Excel file generated ({file_size:.2f} KB)")
        except Exception as e:
            logging.error(f"Excel generation failed: {str(e)}")
            return False

        # 3. Upload the file
        try:
            upload_url = f"{self.client.base_url.rstrip('/')}/api/process-groups/{process_group_id}/sampling-drafts/import"
            logging.info(f"Preparing upload to: {upload_url}")

            # Get token
            token = self.client.token or self.client.get_httpx_client().headers.get(
                "Authorization", ""
            ).replace("Bearer ", "")
            if not token:
                logging.error("No authentication token available")
                return False

            headers = {"Authorization": f"Bearer {token}"}
            data = {"processGroupId": (None, str(process_group_id))}
            files = {
                "importFile": (
                    "samplingplan.xlsx",
                    excel_stream,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            }

            logging.info("Starting file upload...")
            response = requests.post(
                upload_url, headers=headers, data=data, files=files
            )

            if response.ok:
                logging.info(f"Upload successful! Response: {response.status_code}")
                return True
            else:
                logging.error(f"Upload failed. Status: {response.status_code}")
                logging.error(f"Error response: {response.text}")
                return False
        except Exception as e:
            logging.error(f"Upload error: {str(e)}")
            return False

    def generate_and_upload_sampling_plan_single_process(
        self, process_id: int, number_of_samples: int, sample_prefix: str = "Day"
    ) -> bool:
        """
        Generates and uploads a sampling plan to the specified process ID.

        Args:
            process_id: The ID of the process to upload to
            number_of_samples: Number of samples to generate
            sample_prefix: Prefix for the sampling schedule (e.g., "Day")

        Returns:
            bool: True if upload was successful, False otherwise
        """
        # Validate input
        if not isinstance(process_id, int) or process_id <= 0:
            logging.error(f"Invalid process ID provided: {process_id}")
            return False

        logging.info(f"Starting sampling plan generation for process {process_id}")

        # 1. Fetch the process details
        try:
            endpoint = f"{self.client.base_url.rstrip('/')}/api/processes/{process_id}"
            logging.info(f"Fetching process details from: {endpoint}")
            response = self.client.get_httpx_client().get(endpoint)

            if response.status_code != 200:
                logging.error(
                    f"Failed to fetch process details. Status: {response.status_code}"
                )
                logging.error(f"Response: {response.text}")
                return False

            process = response.json()
            if "name" not in process:
                logging.error("Process data doesn't contain 'name' field")
                return False

        except Exception as e:
            logging.error(f"Error fetching process details: {str(e)}")
            return False

        # 2. Generate Excel file
        try:
            logging.info("Generating Excel file")
            wb = Workbook()
            ws = wb.active

            # Set headers
            headers = ["Process", "Sampling schedule", "External ID"]
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                ws[f"{col_letter}1"] = header
                ws[f"{col_letter}1"].font = Font(bold=True)

            # Populate data
            current_row = 2
            process_name = process.get("name", "Unnamed Process")
            logging.info(
                f"Adding {number_of_samples} samples for process: {process_name}"
            )

            for i in range(1, number_of_samples + 1):
                ws[f"A{current_row}"] = process_name
                ws[f"B{current_row}"] = f"{sample_prefix} {i}"
                ws[f"C{current_row}"] = f"{process_name}_{i:03d}"
                current_row += 1

            # Auto-adjust columns
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column].width = (max_length + 2) * 1.2

            # Save to memory
            excel_stream = BytesIO()
            wb.save(excel_stream)
            excel_stream.seek(0)
            file_size = len(excel_stream.getvalue()) / 1024  # KB
            logging.info(f"Excel file generated ({file_size:.2f} KB)")
        except Exception as e:
            logging.error(f"Excel generation failed: {str(e)}")
            return False

        # 3. Upload the file
        try:
            upload_url = f"{self.client.base_url.rstrip('/')}/api/processes/{process_id}/sampling-drafts/import"
            logging.info(f"Preparing upload to: {upload_url}")

            # Get token
            token = self.client.token or self.client.get_httpx_client().headers.get(
                "Authorization", ""
            ).replace("Bearer ", "")
            if not token:
                logging.error("No authentication token available")
                return False

            headers = {"Authorization": f"Bearer {token}"}
            data = {"processId": (None, str(process_id))}
            files = {
                "importFile": (
                    "samplingplan.xlsx",
                    excel_stream,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            }

            logging.info("Starting file upload...")
            response = requests.post(
                upload_url, headers=headers, data=data, files=files
            )

            if response.ok:
                logging.info(f"Upload successful! Response: {response.status_code}")
                return True
            else:
                logging.error(f"Upload failed. Status: {response.status_code}")
                logging.error(f"Error response: {response.text}")
                return False
        except Exception as e:
            logging.error(f"Upload error: {str(e)}")
            return False

    def upload_file(self, file_path: str) -> Optional[Dict[str, Any]]:
        """
        Uploads a file (typically a ZIP archive for an advanced model) to the stored-files endpoint.
        Returns a dict with file details if successful, else None.
        """
        base_url = self.client._base_url  # Ensure this base URL ends with a slash
        endpoint = f"{base_url}api/stored-files"  # For example: "https://master.qub-lab.io/api/stored-files"
        try:
            with open(file_path, "rb") as f:
                # Use "storedFile" as the key since the backend expects that field name.
                files = {
                    "storedFile": (
                        os.path.basename(file_path),
                        f,
                        "application/zip",  # adjust MIME type if needed
                    )
                }
                response = self.client.get_httpx_client().post(endpoint, files=files)
            if response.status_code == HTTPStatus.OK:
                file_info = response.json()
                logging.info("File uploaded successfully: %s", file_info)
                return file_info
            else:
                logging.error(
                    "File upload failed. HTTP Status: %s, Response: %s",
                    response.status_code,
                    response.text,
                )
                return None
        except Exception as e:
            logging.error("Error during file upload: %s", e)
            return None

    def create_advanced_model(
        self, model_data: dict, file_path: str
    ) -> Optional[Dict[str, Any]]:
        """
        Create a new advanced computable model by first uploading an external file (typically a ZIP archive)
        and then building the model DTO with the provided inputs, outputs, and additional model data.
        Advanced models require an engine type of EXTERNAL_PYTHON.
        """
        try:
            # ------------------------------------------------------
            # Step 1: Upload the external file.
            # ------------------------------------------------------
            file_response = self.upload_file(file_path)
            if not file_response:
                logging.error("Advanced model creation aborted: File upload failed.")
                return None

            # Instead of wrapping the file response in a custom class,
            # convert it to a StoredFileDto instance using its from_dict() method
            model_data["externalFile"] = StoredFileDto.from_dict(file_response)

            # ------------------------------------------------------
            # Step 2: Basic cleanup & mandatory checks.
            # ------------------------------------------------------
            if "engineType" not in model_data:
                raise ValueError("Missing 'engineType' in model data.")
            for ro_field in ["id", "creationDate", "updateDate"]:
                model_data.pop(ro_field, None)
            if model_data.get("sensorType") is None:
                model_data.pop("sensorType", None)

            # ------------------------------------------------------
            # Step 3: Handle physical quantities.
            # ------------------------------------------------------
            model_data = self.handle_physical_quantities(model_data)
            if not model_data:
                raise ValueError("Failed to process physical quantities.")

            # ------------------------------------------------------
            # Step 4: Build sub-dtos for inputs.
            # ------------------------------------------------------
            input_dto_list = []
            for inp in model_data.get("inputs", []):
                pq_unit_dict = inp.get("physicalQuantityUnit", {})
                raw_pq_status = pq_unit_dict.get("status", UNSET)
                if raw_pq_status not in [UNSET, None]:
                    raw_pq_status = PhysicalQuantityUnitDtoStatus(raw_pq_status)
                pq_unit_dto = PhysicalQuantityUnitDto(
                    unit=pq_unit_dict.get("unit", ""),
                    name=pq_unit_dict.get("name", ""),
                    id=pq_unit_dict.get("id", UNSET),
                    physical_quantity_id=pq_unit_dict.get("physicalQuantityId", UNSET),
                    status=raw_pq_status,
                )
                input_dto = ComputableModelInputDto(
                    name=inp.get("name", ""),
                    physical_quantity_unit=pq_unit_dto,
                    id=inp.get("id", UNSET),
                    description=inp.get("description", UNSET),
                    order=inp.get("order", UNSET),
                )
                input_dto_list.append(input_dto)

            # ------------------------------------------------------
            # Step 5: Build sub-dtos for outputs.
            # ------------------------------------------------------
            output_dto_list = []
            for outp in model_data.get("outputs", []):
                if outp.get("publishField") is None:
                    outp.pop("publishField", None)
                pq_unit_dict = outp.get("physicalQuantityUnit", {})
                raw_pq_status = pq_unit_dict.get("status", UNSET)
                if raw_pq_status not in [UNSET, None]:
                    raw_pq_status = PhysicalQuantityUnitDtoStatus(raw_pq_status)
                pq_unit_dto = PhysicalQuantityUnitDto(
                    unit=pq_unit_dict.get("unit", ""),
                    name=pq_unit_dict.get("name", ""),
                    id=pq_unit_dict.get("id", UNSET),
                    physical_quantity_id=pq_unit_dict.get("physicalQuantityId", UNSET),
                    status=raw_pq_status,
                )
                output_dto = ComputableModelOutputDto(
                    name=outp.get("name", ""),
                    physical_quantity_unit=pq_unit_dto,
                    id=outp.get("id", UNSET),
                    description=outp.get("description", UNSET),
                    order=outp.get("order", UNSET),
                    publish_field=UNSET,
                )
                output_dto_list.append(output_dto)

            # ------------------------------------------------------
            # Step 6: Convert top-level strings into Enums and build final DTO.
            # ------------------------------------------------------
            engine_type_str = model_data["engineType"]
            engine_type_enum = AbstractComputableModelDtoEngineType(engine_type_str)
            if engine_type_enum == AbstractComputableModelDtoEngineType.JYTHON:
                logging.error(
                    "Advanced model creation requires 'EXTERNAL_PYTHON' engine type."
                )
                return None
            calc_style_str = model_data.get("calculationStyle", "ONLINE")
            calc_style_enum = AbstractComputableModelDtoCalculationStyle(calc_style_str)
            status_str = model_data.get("status", "DRAFT")
            status_enum = AbstractComputableModelDtoStatus(status_str)

            # Note: The 'script' parameter is removed for advanced models.
            model_body = ExternalPythonComputableModelDto(
                engine_type=engine_type_enum,
                kpi_name=model_data.get("kpiName", ""),
                abbr=model_data.get("abbr", ""),
                calculation_style=calc_style_enum,
                outputs=output_dto_list,
                inputs=input_dto_list,
                status=status_enum,
                description=model_data.get("description", UNSET),
                external_file=model_data.get("externalFile", UNSET),
            )

            # ------------------------------------------------------
            # Step 7: Call the API.
            # ------------------------------------------------------
            response = create_computable_model.sync_detailed(
                client=self.client,
                body=model_body,
            )

            if response.status_code == HTTPStatus.OK:
                parsed = json.loads(response.content.decode())
                logging.info("Successfully created advanced computable model.")
                return parsed
            elif response.status_code == HTTPStatus.BAD_REQUEST:
                if "valid.unique.kpi.name" in json.loads(response.content.decode()):
                    logging.error(
                        "Failed to create model because 'kpiName' is already in use. Please choose a different kpiName."
                    )
                else:
                    logging.error(
                        "Failed to create advanced computable model. Response: %s",
                        response.content.decode(),
                    )
                return None
            else:
                logging.error(
                    "Failed to create advanced computable model. HTTP Status: %s. Response: %s",
                    response.status_code,
                    response.content.decode(),
                )
                return None

        except Exception as e:
            logging.error("Error creating advanced computable model: %s", e)
            return None


# Internal function (not for public use)
def _internal_function():
    pass


# Dynamically create __all__
__all__ = [name for name in globals() if not name.startswith("_")]
