import json
import openpyxl
import os
from typing import Dict, Any
from .ikea_status import task_status_to_str
from typing import Dict, Any, Optional
from xianbao.core.context import global_context


class ExcelWriter:
    def __init__(self, xlsx_path: str, db_path: str=r"D:\install\ylsw\ums-rpa\rpa_ztdb.db", template_path: str = 'E:\\template.xlsx'):
        self.template_path = template_path
        self.ref_cache = {}  # 参考号缓存字典
        self._rpa_fill = RpaFillCondition(db_path=db_path)
        self.xlsx_path = xlsx_path
        self.db_path = db_path

        # 检查文件是否存在，不存在则使用模板
        if not os.path.exists(xlsx_path):
            self.workbook = openpyxl.load_workbook(self.template_path)
        else:
            self.workbook = openpyxl.load_workbook(xlsx_path)
            
        # 初始化或更新参考号缓存
        self._init_ref_cache(self.workbook.active)
        self.sheet = self.workbook.active

    def _init_ref_cache(self, sheet):
        """初始化参考号缓存"""
        self.ref_cache.clear()
        for idx, cell in enumerate(sheet['A'][1:], start=5):  # B列是参考号列，从第2行开始
            if cell.value:
                self.ref_cache[cell.value] = idx
    
    def _find_row_by_ref(self, ref_no: str) -> Optional[int]:
        """查找参考号对应的行号"""
        # 优先从缓存查找
        if ref_no in self.ref_cache:
            return self.ref_cache[ref_no]
        return None
        
    def write_to_excel(self, task: Dict[str, Any]) -> bool:
        """
        将字典数据写入Excel文件
        :param data: 输入字典数据
        :param xlsx_path: xlsx文件路径
        :return: 是否成功
        """
        try:
            meta = task['meta']
            if meta is None or meta == '':
                return False
            task['meta'] = meta if isinstance(meta, dict) else json.loads(meta)
            meta = task['meta']
            # 查找参考号对应的行
            task_id = task['id']
            row_idx = self._find_row_by_ref(task_id) if task_id else None
            
            # 准备数据映射
            row_data = self._rpa_fill._data_mapper(task)
            
            # 写入数据
            if row_idx is not None:
                # 更新现有行
                for col_idx, value in enumerate(row_data, 1):
                    self.sheet.cell(row=row_idx, column=col_idx, value=value)
            elif row_data is not None:
                # 新增行
                self.sheet.append(list(row_data))
            return True
        except Exception as e:
            print(f"写入Excel失败: {str(e)}")
            return False

    def close(self):
        """关闭Excel文件"""
        self.workbook.save(self.xlsx_path)
        self.workbook.close()



class RpaFillCondition:
    def __init__(self, db_path: str):
        self.sys_config = global_context['dict_manager'].get_dict_as_mapping('sys_config')

    # 默认的数据映射实现
    def _data_mapper(self, task: Dict[str, Any]) -> Dict[str, Any]:
        """
        默认的数据映射实现
        :param task: 输入字典数据
        :return: 映射后的字典数据
        """
        data = task['meta']
        log = task['log']
        if log is not None and log != '':
            log += '\n'
        else:
            log += task_status_to_str(task['status'])

        res = [
            # 序号: 自动生成
            task['id'],
            # 参考号
            data.get('ums_receipt_info', {}).get('参考号', ''),
            # 异常描述
            log,
            # 电脑文件存储地址
            os.path.join(self.sys_config['workpath'], data.get('relative_path', '')),
            # 网盘地址
            '',
            # （必填）签购单
            data.get('ums_receipt', ''),
            # （必填）商品照片
            data.get('ikea_receipt', ''),
            # （必填）自提单或送货单
            data.get('delivery_note', ''),
            # 发票图片名称
            data.get('invoice', ''),
            # 发票代码
            data.get('invoice_info', {}).get('发票号码', ''),
            # 发票金额
            data.get('ikea_receipt_info', {}).get('合计', ''),
            # 销售企业名称（注册名称）
            '',
            # 统一社会信用代码
            data.get('invoice_info', {}).get('购买方统一社会信用代码/纳税人识别号', ''),
            # 购买方名称
            data.get('invoice_info', {}).get('购买方名称', ''),
            # 销售企业注册地区行政区划
            self.sys_config['销售企业注册地区行政区划'],
            # 销售企业注册详细地址
            self.sys_config['销售企业注册详细地址'],
            # 购买方姓名
            data.get('invoice_info', {}).get('购买方名称', ''),
            # 手机号
            data.get('delivery_info', {}).get('电话', ''),
            # 所在地区
            data.get('delivery_info', {}).get('地区', ''),
            # 详细地址
            data.get('delivery_info', {}).get('地址', ''),
            # 配送情况
            '配送',
            # 送货（提货）时间
            data.get('ums_receipt_info', {}).get('日期时间', '')
        ]
        return res

