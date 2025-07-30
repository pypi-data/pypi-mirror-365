import pandas as pd
import json
import openpyxl
import os
import re
from typing import Dict, Any
from .ikea_status import task_status_to_str
from typing import Dict, Any, Optional


class ExcelWriter:
    def __init__(self, xlsx_path: str, order_path: str, template_path: str = 'E:\\template.xlsx'):
        self.template_path = template_path
        self.ref_cache = {}  # 参考号缓存字典
        self._rpa_fill = RpaBackfillingRecong(order_path)
        self.xlsx_path = xlsx_path

        # 检查文件是否存在，不存在则使用模板
        if not os.path.exists(xlsx_path):
            self.workbook = openpyxl.load_workbook(self.template_path)
        else:
            self.workbook = openpyxl.load_workbook(xlsx_path)
            
        # 初始化或更新参考号缓存
        self._init_ref_cache(self.workbook.active)
        self.sheet = self.workbook.active

    def _init_ref_cache(self, sheet):
        """初始化参考号缓存"""
        self.ref_cache.clear()
        for idx, cell in enumerate(sheet['B'][1:], start=5):  # B列是参考号列，从第5行开始
            if cell.value:
                self.ref_cache[cell.value] = idx
    
    def _find_row_by_ref(self, ref_no: str) -> Optional[int]:
        """查找参考号对应的行号"""
        # 优先从缓存查找
        if ref_no in self.ref_cache:
            return self.ref_cache[ref_no]
        return None
        
    def write_to_excel(self, task: Dict[str, Any]) -> bool:
        """
        将字典数据写入Excel文件
        :param data: 输入字典数据
        :param xlsx_path: xlsx文件路径
        :return: 是否成功
        """
        try:
            meta = task['meta']
            if meta is None or meta == '':
                return False
            task['meta'] = meta if isinstance(meta, dict) else json.loads(meta)
            meta = task['meta']
            # 查找参考号对应的行
            ref_no = meta.get('ums_receipt_info', {}).get('参考号', '')
            row_idx = self._find_row_by_ref(ref_no) if ref_no else None
            
            # 准备数据映射
            row_data = self._rpa_fill._data_mapper(task)
            
            # 写入数据
            if row_idx is not None:
                # 更新现有行
                for col_idx, value in enumerate(row_data, 1):
                    self.sheet.cell(row=row_idx, column=col_idx, value=value)
            elif row_data is not None:
                # 新增行
                self.sheet.append(list(row_data))
            return True
        except Exception as e:
            print(f"写入Excel失败: {str(e)}")
            return False

    def close(self):
        """关闭Excel文件"""
        self.workbook.save(self.xlsx_path)
        self.workbook.close()


class RpaBackfillingRecong:
    def __init__(self, order_path: str):
        # 预加载订单表
        self.order_df = None
        order_file = order_path
        if os.path.exists(order_file):
            self.order_df = pd.read_excel(order_file, dtype=str)
            # 转换日期格式
            if 'transdate' in self.order_df.columns:
                self.order_df['transdate'] = pd.to_datetime(self.order_df['transdate'], format='%Y%m%d').dt.strftime('%Y-%m-%d')
            else:
                self.order_df['transdate'] = ''
    
    def find_22_digits(self, text):
        """
        从字符串中识别22位连续数字
        :param text: 输入字符串
        :return: 匹配到的22位数字字符串，如果没有则返回None
        """
        pattern = r'\d{22}'
        match = re.search(pattern, text)
        return match.group(0) if match else None


    def _data_mapper(self, task: Dict[str, Any]) -> Dict[str, Any]:
        log = ''
        if self.order_df is None:
            log = "订单表未加载，请检查订单表.xlsx文件是否存在\n"
            raise ValueError(log)

        meta = task['meta']
        
        # 匹配参考号
        ref_no = meta.get('ums_receipt_info', {}).get('参考号', '')
        matched_row = self.order_df[self.order_df['traceno'] == ref_no]
        
        if matched_row.empty:
            return None
            
        # 提取匹配的行数据
        transdate = matched_row['transdate'].values[0]
        order_amount = matched_row['order_amount'].values[0]
        should_amount = matched_row['should_amount'].values[0]
        chl_discount = matched_row['chl_discount'].values[0]
        termid = matched_row['termid'].values[0]

        if log is not None and log != '':
            log += '\n'
        else:
            log += task_status_to_str(task['status'])
        
        # 返回指定格式的数组
        return [
            # '交易时间': 交易发生的时间
            transdate,
            # '参考号': 交易参考编号
            ref_no,
            # '终端号': 交易终端编号
            termid,
            # '顾客姓名': 顾客姓名
            meta.get('invoice_info', {}).get('购买方名称', ''),
            # '交易金额': 交易总金额
            order_amount,
            # '顾客实付金额': 顾客实际支付金额
            should_amount,
            # '政府补贴金额': 政府补贴金额
            chl_discount,
            # '已回款金额': 已回款金额
            None,
            # '状态': 交易状态
            log,
            # '宜家发票号': 宜家发票号码
            meta.get('invoice_info', {}).get('发票号码', ''),
            # '含税发票金额': 发票含税金额
            meta.get('invoice_info', {}).get('合计', ''),
            # '宜家订单号': 宜家订单编号
            self.find_22_digits(meta.get('invoice_info', {}).get('备注', '')),
            # '宜家收银小票含税金额': 收银小票含税金额
            None,
            # '宜家产品金额': 产品金额
            None,
            # '宜家服务费': 服务费金额
            None,
            # '宜家SPA/Coupon金额': SPA/Coupon金额
            None,
            # '差额(发票与银联交易)': 发票与银联交易差额
            None,
            # '差额(发票与宜家交易)': 发票与宜家交易差额
            None,
            # '差额（银联交易-宜家交易）': 银联与宜家交易差额
            None,
            # '回款差额': 回款差额
            None
        ]
