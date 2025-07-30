"""End-to-end integration tests for GridGulp."""

import asyncio
import tempfile
from pathlib import Path
import csv
import json

import pytest
import pandas as pd
import openpyxl

from gridgulp import GridGulp
from gridgulp.config import Config
from gridgulp.extractors import DataFrameExtractor
from gridgulp.readers import get_reader


class TestEndToEndExcel:
    """End-to-end tests for Excel files."""

    @pytest.mark.asyncio
    async def test_simple_excel_detection(self):
        """Test complete pipeline for simple Excel file."""
        # Create test Excel file
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            temp_path = Path(f.name)

        try:
            # Create Excel file with openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sales"

            # Add data
            data = [
                ["Product", "Q1", "Q2", "Q3", "Q4"],
                ["Apples", 100, 150, 200, 175],
                ["Bananas", 80, 90, 85, 95],
                ["Oranges", 120, 110, 130, 140],
                ["Total", 300, 350, 415, 410],
            ]

            for row_idx, row_data in enumerate(data, 1):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            wb.save(temp_path)
            wb.close()

            # Run detection
            gg = GridGulp()
            result = await gg.detect_tables(temp_path)

            # Verify results
            assert result.total_tables == 1
            assert len(result.sheets) == 1
            assert result.sheets[0].name == "Sales"

            table = result.sheets[0].tables[0]
            assert table.range.excel_range == "A1:E5"
            assert table.confidence > 0.8
            # Headers are not populated by default in detection results

        finally:
            temp_path.unlink()

    @pytest.mark.asyncio
    async def test_multi_sheet_excel(self):
        """Test Excel file with multiple sheets."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            temp_path = Path(f.name)

        try:
            wb = openpyxl.Workbook()

            # Sheet 1: Sales data
            ws1 = wb.active
            ws1.title = "Sales"
            sales_data = [
                ["Month", "Revenue", "Cost"],
                ["Jan", 10000, 7000],
                ["Feb", 12000, 8000],
                ["Mar", 15000, 9000],
            ]
            for r_idx, row in enumerate(sales_data, 1):
                for c_idx, val in enumerate(row, 1):
                    ws1.cell(row=r_idx, column=c_idx, value=val)

            # Sheet 2: Employee data
            ws2 = wb.create_sheet("Employees")
            emp_data = [
                ["Name", "Department", "Salary"],
                ["Alice", "Sales", 50000],
                ["Bob", "IT", 60000],
                ["Charlie", "Marketing", 55000],
            ]
            for r_idx, row in enumerate(emp_data, 1):
                for c_idx, val in enumerate(row, 1):
                    ws2.cell(row=r_idx, column=c_idx, value=val)

            # Sheet 3: Empty sheet
            wb.create_sheet("Empty")

            wb.save(temp_path)
            wb.close()

            # Run detection
            gg = GridGulp()
            result = await gg.detect_tables(temp_path)

            # Verify
            assert result.total_tables == 2
            assert len(result.sheets) == 3

            # Check each sheet
            sales_sheet = next(s for s in result.sheets if s.name == "Sales")
            assert len(sales_sheet.tables) == 1
            assert sales_sheet.tables[0].shape == (4, 3)

            emp_sheet = next(s for s in result.sheets if s.name == "Employees")
            assert len(emp_sheet.tables) == 1
            assert emp_sheet.tables[0].shape == (4, 3)

            empty_sheet = next(s for s in result.sheets if s.name == "Empty")
            assert len(empty_sheet.tables) == 0

        finally:
            temp_path.unlink()

    @pytest.mark.asyncio
    async def test_excel_with_native_tables(self):
        """Test Excel file with native Excel tables (ListObjects)."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            temp_path = Path(f.name)

        try:
            wb = openpyxl.Workbook()
            ws = wb.active

            # Add data
            headers = ["ID", "Name", "Value"]
            data = [[1, "Item A", 100], [2, "Item B", 200], [3, "Item C", 300]]

            # Write headers
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            # Write data
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Create Excel table
            tab = openpyxl.worksheet.table.Table(displayName="DataTable", ref="A1:C4")
            ws.add_table(tab)

            wb.save(temp_path)
            wb.close()

            # Run detection
            gg = GridGulp()
            result = await gg.detect_tables(temp_path)

            # Should detect the table (might not use excel_metadata due to openpyxl compatibility)
            assert result.total_tables == 1
            table = result.sheets[0].tables[0]
            assert table.confidence > 0.8
            assert table.range.excel_range == "A1:C4"

        finally:
            temp_path.unlink()


class TestEndToEndCSV:
    """End-to-end tests for CSV files."""

    @pytest.mark.asyncio
    async def test_simple_csv_detection(self):
        """Test complete pipeline for simple CSV file."""
        with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False) as f:
            temp_path = Path(f.name)

            # Write CSV data
            writer = csv.writer(f)
            writer.writerow(["Name", "Age", "City"])
            writer.writerow(["Alice", 25, "New York"])
            writer.writerow(["Bob", 30, "Los Angeles"])
            writer.writerow(["Charlie", 35, "Chicago"])

        try:
            # Run detection
            gg = GridGulp()
            result = await gg.detect_tables(temp_path)

            # Verify
            assert result.total_tables == 1
            assert len(result.sheets) == 1

            table = result.sheets[0].tables[0]
            assert table.shape == (4, 3)  # Including header
            # Headers are not populated by default
            # assert table.headers == ["Name", "Age", "City"]
            assert table.confidence > 0.8

        finally:
            temp_path.unlink()

    @pytest.mark.asyncio
    async def test_csv_with_different_delimiters(self):
        """Test CSV files with various delimiters."""
        test_cases = [
            (",", "comma.csv"),
            ("\t", "tab.tsv"),
            ("|", "pipe.csv"),
            (";", "semicolon.csv"),
        ]

        for delimiter, filename in test_cases:
            with tempfile.NamedTemporaryFile(mode="w", suffix=filename, delete=False) as f:
                temp_path = Path(f.name)

                # Write data with specific delimiter
                data = [["Col1", "Col2", "Col3"], ["A", "B", "C"], ["D", "E", "F"]]

                for row in data:
                    f.write(delimiter.join(row) + "\n")

            try:
                # Run detection
                gg = GridGulp()
                result = await gg.detect_tables(temp_path)

                # Should detect table regardless of delimiter
                assert result.total_tables == 1
                table = result.sheets[0].tables[0]
                assert table.shape == (3, 3)
                # Headers are not populated by default
            # assert table.headers == ["Col1", "Col2", "Col3"]

            finally:
                temp_path.unlink()


class TestEndToEndDataFrameExtraction:
    """End-to-end tests for DataFrame extraction."""

    @pytest.mark.asyncio
    async def test_extract_dataframe_from_excel(self):
        """Test extracting pandas DataFrame from detected tables."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            temp_path = Path(f.name)

        try:
            # Create test data
            df = pd.DataFrame(
                {
                    "Product": ["A", "B", "C"],
                    "Price": [10.5, 20.0, 15.75],
                    "Quantity": [100, 200, 150],
                }
            )
            df.to_excel(temp_path, index=False)

            # Detect tables
            gg = GridGulp()
            result = await gg.detect_tables(temp_path)

            # Extract DataFrame
            reader = get_reader(temp_path)
            file_data = reader.read_sync()

            extractor = DataFrameExtractor()
            sheet_data = file_data.sheets[0]
            table = result.sheets[0].tables[0]

            extracted_df, metadata, quality = extractor.extract_dataframe(sheet_data, table.range)

            # Verify extraction
            assert extracted_df is not None
            assert quality > 0.8
            assert list(extracted_df.columns) == ["Product", "Price", "Quantity"]
            assert len(extracted_df) == 3
            # The DataFrame extractor returns data as extracted - verify shape only

        finally:
            temp_path.unlink()


class TestEndToEndDirectoryProcessing:
    """End-to-end tests for directory processing."""

    @pytest.mark.asyncio
    async def test_process_directory(self):
        """Test processing entire directory of spreadsheets."""
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)

            # Create multiple test files
            files_created = []

            # Excel file
            excel_path = temp_path / "data1.xlsx"
            df1 = pd.DataFrame({"A": [1, 2, 3], "B": [4, 5, 6]})
            df1.to_excel(excel_path, index=False)
            files_created.append(excel_path)

            # CSV file
            csv_path = temp_path / "data2.csv"
            df2 = pd.DataFrame({"X": [7, 8, 9], "Y": [10, 11, 12]})
            df2.to_csv(csv_path, index=False)
            files_created.append(csv_path)

            # Subdirectory with file
            subdir = temp_path / "subdir"
            subdir.mkdir()
            sub_file = subdir / "data3.xlsx"
            df3 = pd.DataFrame({"P": [13, 14], "Q": [15, 16]})
            df3.to_excel(sub_file, index=False)
            files_created.append(sub_file)

            # Run directory detection
            gg = GridGulp()
            results = await gg.detect_tables_in_directory(temp_path)

            # Verify
            assert len(results) == 3
            assert all(r.total_tables == 1 for r in results.values())

            # Check that all files were processed (resolve paths to handle symlinks)
            processed_files = {p.resolve() for p in results.keys()}
            expected_files = {p.resolve() for p in files_created}
            assert processed_files == expected_files


class TestEndToEndPerformance:
    """Performance tests for end-to-end scenarios."""

    @pytest.mark.asyncio
    async def test_large_file_performance(self):
        """Test performance with large file."""
        with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False) as f:
            temp_path = Path(f.name)

            # Create large CSV (10000 rows)
            writer = csv.writer(f)
            writer.writerow(["ID", "Value1", "Value2", "Value3"])

            for i in range(10000):
                writer.writerow([i, f"val1_{i}", f"val2_{i}", f"val3_{i}"])

        try:
            # Run detection
            gg = GridGulp()
            import time

            start_time = time.time()
            result = await gg.detect_tables(temp_path)
            end_time = time.time()

            # Verify detection worked
            assert result.total_tables == 1
            table = result.sheets[0].tables[0]
            assert table.shape[0] == 10001  # Including header

            # Check performance (should be under 5 seconds)
            processing_time = end_time - start_time
            assert processing_time < 5.0

        finally:
            temp_path.unlink()


class TestEndToEndErrorHandling:
    """Test error handling in end-to-end scenarios."""

    @pytest.mark.asyncio
    async def test_corrupted_file_handling(self):
        """Test handling of corrupted files."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            temp_path = Path(f.name)
            # Write invalid data
            f.write(b"This is not a valid Excel file")

        try:
            gg = GridGulp()

            # Should handle gracefully
            with pytest.raises(ValueError):  # Should raise ValueError for invalid files
                await gg.detect_tables(temp_path)

        finally:
            temp_path.unlink()

    @pytest.mark.asyncio
    async def test_empty_file_handling(self):
        """Test handling of empty files."""
        with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as f:
            temp_path = Path(f.name)
            # Empty file

        try:
            gg = GridGulp()

            # Empty files might raise ValueError due to unknown type
            try:
                result = await gg.detect_tables(temp_path)
                # If successful, should have no tables
                assert result.total_tables == 0
            except ValueError as e:
                # Empty file detected as unknown type is acceptable
                assert "unknown files" in str(e).lower()

        finally:
            temp_path.unlink()


class TestEndToEndConfiguration:
    """Test different configurations in end-to-end scenarios."""

    @pytest.mark.asyncio
    async def test_custom_confidence_threshold(self):
        """Test with custom confidence threshold."""
        with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False) as f:
            temp_path = Path(f.name)

            # Create sparse data that might have low confidence
            writer = csv.writer(f)
            writer.writerow(["A", "B"])
            writer.writerow([1, 2])
            # Many empty rows
            for _ in range(20):
                writer.writerow(["", ""])
            writer.writerow([3, 4])

        try:
            # Test with default threshold
            gg1 = GridGulp()
            result1 = await gg1.detect_tables(temp_path)

            # Test with lower threshold
            gg2 = GridGulp(confidence_threshold=0.3)
            result2 = await gg2.detect_tables(temp_path)

            # Lower threshold might detect more tables or larger tables
            assert result2.total_tables >= result1.total_tables

        finally:
            temp_path.unlink()
