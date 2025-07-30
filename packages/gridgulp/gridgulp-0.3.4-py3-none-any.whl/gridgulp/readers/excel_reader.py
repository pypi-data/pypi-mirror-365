"""Excel file reader supporting both modern and legacy formats."""

import contextlib
import logging
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from ..detectors.excel_metadata_extractor import ExcelMetadata, ExcelMetadataExtractor
from ..models.file_info import FileInfo, FileType
from ..models.sheet_data import CellData, FileData, SheetData
from .base_reader import (
    CorruptedFileError,
    PasswordProtectedError,
    ReaderError,
    SyncBaseReader,
)

logger = logging.getLogger(__name__)

# Optional imports - cached at module level
try:
    import olefile

    HAS_OLEFILE = True
except ImportError:
    HAS_OLEFILE = False

try:
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import xlrd
    from xlrd import XLRDError

    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

    # Define a placeholder exception to prevent NameError
    XLRDError = type("XLRDError", (Exception,), {})


class ExcelReader(SyncBaseReader):
    """Reader for Excel files (.xlsx, .xlsm, .xls).

    This reader handles both modern Excel formats (XLSX/XLSM) using openpyxl
    and legacy formats (XLS) using xlrd. It automatically selects the appropriate
    backend based on file type and extracts cell data, formatting, and metadata.

    Examples
    --------
    Basic usage::

        >>> from gridgulp.readers import get_reader
        >>> reader = get_reader("sales_report.xlsx")
        >>> file_data = reader.read_sync()
        >>> print(f"Sheets: {len(file_data.sheets)}")

    With context manager::

        >>> with ExcelReader(path, file_info) as reader:
        ...     file_data = reader.read_sync()
        ...     # Workbook is automatically closed

    Notes
    -----
    The reader preserves formatting information including borders, bold/italic
    text, and cell alignment. Excel metadata such as ListObjects (tables) and
    named ranges are also extracted when available.

    Password-protected files are detected and raise appropriate errors. XLSB
    format is detected but not supported and will raise UnsupportedFormatError.
    """

    def __init__(self, file_path: Path, file_info: FileInfo):
        """Initialize Excel reader.

        Args
        ----
        file_path : Path
            Path to the Excel file to read.
        file_info : FileInfo
            Pre-analyzed file information including detected type.

        Notes
        -----
        The appropriate backend (openpyxl or xlrd) is selected based on file type:
        - XLSX/XLSM: Uses openpyxl for full feature support
        - XLS: Uses xlrd for legacy format compatibility
        """
        super().__init__(file_path, file_info)
        self._workbook: Any = None  # Type varies by backend
        self._use_openpyxl = file_info.type in {
            FileType.XLSX,
            FileType.XLSM,
        }
        self._metadata_extractor = ExcelMetadataExtractor()
        self._excel_metadata: ExcelMetadata | None = None

    def can_read(self) -> bool:
        """Check if can read Excel files.

        Returns
        -------
        bool
            True if the file type is XLSX, XLS, or XLSM; False otherwise.

        Notes
        -----
        XLSB format is explicitly not supported even though it's an Excel format.
        This method is used by the factory to validate reader compatibility.
        """
        return self.file_info.type in {
            FileType.XLSX,
            FileType.XLS,
            FileType.XLSM,
        }

    def get_supported_formats(self) -> list[str]:
        """Get supported Excel formats.

        Returns
        -------
        list[str]
            List of supported file extensions: ["xlsx", "xls", "xlsm"].

        Notes
        -----
        This is a static list of formats this reader can handle. XLSB is
        intentionally excluded as it requires specialized parsing not
        currently implemented.
        """
        return ["xlsx", "xls", "xlsm"]

    def __enter__(self) -> "ExcelReader":
        """Context manager entry.

        Returns
        -------
        ExcelReader
            Returns self for use in with statements.
        """
        return self

    def __exit__(self, exc_type: Any, exc_val: Any, exc_tb: Any) -> None:
        """Context manager exit - cleanup resources.

        Args
        ----
        exc_type : Any
            Exception type if an exception occurred.
        exc_val : Any
            Exception value if an exception occurred.
        exc_tb : Any
            Exception traceback if an exception occurred.

        Notes
        -----
        Always closes the workbook regardless of whether an exception occurred.
        This ensures Excel file handles are properly released.
        """
        self.close()

    def close(self) -> None:
        """Close workbook and free resources.

        Notes
        -----
        This method safely closes the Excel workbook and releases file handles.
        It's automatically called when using the reader as a context manager,
        but can also be called manually. Multiple calls are safe - subsequent
        calls have no effect.
        """
        if self._workbook is not None:
            try:
                if hasattr(self._workbook, "close"):
                    self._workbook.close()
            except Exception as e:
                self.logger.debug(f"Error closing workbook: {e}")
            finally:
                self._workbook = None

    def _is_password_protected(self) -> bool:
        """Check if Excel file is password protected.

        Password-protected Excel files are stored as OLE containers
        with EncryptedPackage and EncryptionInfo streams.
        """
        if not HAS_OLEFILE:
            # If olefile is not available, we can't detect password protection
            # Let it fail with the normal error later
            return False

        try:
            if olefile.isOleFile(str(self.file_path)):
                with olefile.OleFileIO(str(self.file_path)) as ole:
                    streams = ole.listdir()
                    # Check for encryption markers
                    for stream in streams:
                        stream_str = "/".join(stream) if isinstance(stream, list) else str(stream)
                        if "EncryptedPackage" in stream_str or "EncryptionInfo" in stream_str:
                            return True
            return False
        except Exception:
            # Any error in detection, assume not password protected
            return False

    def read_sync(self) -> FileData:
        """Read Excel file synchronously.

        Returns:
            FileData with all sheets

        Raises:
            ReaderError: If file cannot be read
        """
        self._log_read_start()
        self.validate_file()

        # Check for password protection first
        if self._is_password_protected():
            raise PasswordProtectedError(f"File is password protected: {self.file_path}")

        try:
            if self._use_openpyxl:
                return self._read_with_openpyxl()
            else:
                return self._read_with_xlrd()
        except Exception as e:
            error_msg = str(e).lower()
            if "password" in error_msg:
                raise PasswordProtectedError(f"File is password protected: {self.file_path}") from e
            elif "corrupt" in error_msg or "invalid" in error_msg:
                raise CorruptedFileError(f"File appears to be corrupted: {self.file_path}") from e
            else:
                raise ReaderError(f"Failed to read Excel file: {e}") from e

    def _read_with_openpyxl(self) -> FileData:
        """Read modern Excel file using openpyxl."""
        if not HAS_OPENPYXL:
            raise ReaderError(
                "openpyxl is required to read .xlsx files. "
                "Please install it with: pip install openpyxl"
            )

        try:
            # Load workbook with data_only=False to preserve formulas
            self._workbook = load_workbook(
                filename=str(self.file_path), data_only=False, read_only=False
            )

            sheets = []
            for sheet_name in self._workbook.sheetnames:
                worksheet = self._workbook[sheet_name]
                sheet_data = self._read_openpyxl_sheet(worksheet)
                sheets.append(sheet_data)

            # Extract file metadata
            metadata = self._extract_openpyxl_metadata()

            # Extract Excel-specific metadata
            self._excel_metadata = self._metadata_extractor.extract_metadata_openpyxl(
                self._workbook
            )

            file_data = FileData(
                sheets=sheets,
                metadata=metadata,
                file_format=self.file_info.type.value,
                application=metadata.get("application"),
                version=metadata.get("version"),
            )

            self._log_read_complete(file_data)
            return file_data

        except InvalidFileException as e:
            self.close()  # Ensure cleanup on error
            raise CorruptedFileError(f"Invalid Excel file format: {e}") from e
        except Exception as e:
            self.close()  # Ensure cleanup on error
            raise ReaderError(f"Failed to read with openpyxl: {e}") from e
        finally:
            # Always close the workbook after reading
            self.close()

    def _read_openpyxl_sheet(self, worksheet: Any) -> SheetData:
        """Read a single sheet using openpyxl.

        Args:
            worksheet: openpyxl worksheet object

        Returns:
            SheetData with all cell information
        """
        sheet_data = SheetData(name=worksheet.title, is_visible=worksheet.sheet_state == "visible")

        # Iterate through all cells with data
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None or cell.has_style:
                    cell_data = self._convert_openpyxl_cell(cell)
                    if cell_data:
                        sheet_data.set_cell(cell.row - 1, cell.column - 1, cell_data)

        return sheet_data

    def _convert_openpyxl_cell(self, cell: Any) -> CellData | None:
        """Convert openpyxl cell to CellData.

        Args:
            cell: openpyxl cell object

        Returns:
            CellData object or None if cell should be skipped
        """
        try:
            # Get cell value and handle different types
            value = cell.value
            if isinstance(value, datetime):
                # Keep datetime objects as-is
                pass
            elif hasattr(value, "__float__"):
                # Convert to Decimal for precise numbers
                with contextlib.suppress(Exception):
                    value = Decimal(str(value))

            # Determine data type
            data_type = self._get_data_type(value)

            # Extract formatting information
            font = cell.font
            fill = cell.fill
            alignment = cell.alignment

            # Handle merged cells
            is_merged = False
            merge_range = None
            if hasattr(cell.parent, "merged_cells"):
                for merged_range_obj in cell.parent.merged_cells.ranges:
                    if cell.coordinate in merged_range_obj:
                        is_merged = True
                        merge_range = str(merged_range_obj)
                        break

            # Extract indentation level and alignment
            indentation_level = 0
            horizontal_alignment = None
            if alignment:
                # openpyxl stores indent as an integer (but sometimes returns float)
                indentation_level = int(alignment.indent) if alignment.indent is not None else 0
                # Get horizontal alignment
                horizontal_alignment = alignment.horizontal

            # Extract border information
            border = cell.border
            border_top = self._get_border_style(border.top) if border else None
            border_bottom = self._get_border_style(border.bottom) if border else None
            border_left = self._get_border_style(border.left) if border else None
            border_right = self._get_border_style(border.right) if border else None

            return CellData(
                value=value,
                formatted_value=str(value) if value is not None else None,
                data_type=data_type,
                is_bold=font.bold if font.bold is not None else False,
                is_italic=font.italic if font.italic is not None else False,
                is_underline=font.underline is not None and font.underline != "none",
                font_size=font.size,
                font_color=self._get_color_hex(font.color),
                background_color=self._get_fill_color_hex(fill),
                border_top=border_top,
                border_bottom=border_bottom,
                border_left=border_left,
                border_right=border_right,
                is_merged=is_merged,
                merge_range=merge_range,
                has_formula=cell.data_type == "f",
                formula=cell.value if cell.data_type == "f" else None,
                indentation_level=indentation_level,
                alignment=horizontal_alignment,
                row=cell.row - 1,  # Convert to 0-based
                column=cell.column - 1,  # Convert to 0-based
            )

        except Exception as e:
            self.logger.warning(f"Failed to convert cell {cell.coordinate}: {e}")
            return None

    def _read_with_xlrd(self) -> FileData:
        """Read legacy Excel file using xlrd."""
        if not HAS_XLRD:
            raise ReaderError(
                "xlrd is required to read .xls files. Please install it with: pip install xlrd"
            )

        try:
            self._workbook = xlrd.open_workbook(str(self.file_path), formatting_info=True)

            sheets = []
            for sheet_idx in range(self._workbook.nsheets):
                worksheet = self._workbook.sheet_by_index(sheet_idx)
                sheet_data = self._read_xlrd_sheet(worksheet)
                sheets.append(sheet_data)

            # Extract metadata
            metadata = self._extract_xlrd_metadata()

            # Extract Excel-specific metadata (limited for xlrd)
            self._excel_metadata = self._metadata_extractor.extract_metadata_xlrd(self._workbook)

            file_data = FileData(
                sheets=sheets,
                metadata=metadata,
                file_format="xls",
                application="Microsoft Excel",
            )

            self._log_read_complete(file_data)
            return file_data

        except XLRDError as e:
            self.close()  # Ensure cleanup on error
            raise CorruptedFileError(f"Invalid XLS file: {e}") from e
        except Exception as e:
            self.close()  # Ensure cleanup on error
            raise ReaderError(f"Failed to read with xlrd: {e}") from e
        finally:
            # Always close the workbook after reading
            self.close()

    def _read_xlrd_sheet(self, worksheet: Any) -> SheetData:
        """Read a single sheet using xlrd.

        Args:
            worksheet: xlrd sheet object

        Returns:
            SheetData with all cell information
        """
        sheet_data = SheetData(
            name=worksheet.name,
            is_visible=worksheet.visibility == 0,  # 0 = visible in xlrd
        )

        # Iterate through all cells
        for row_idx in range(worksheet.nrows):
            for col_idx in range(worksheet.ncols):
                cell = worksheet.cell(row_idx, col_idx)
                if cell.value != "" or cell.ctype != 0:  # Not empty
                    cell_data = self._convert_xlrd_cell(cell, row_idx, col_idx, worksheet)
                    if cell_data:
                        sheet_data.set_cell(row_idx, col_idx, cell_data)

        return sheet_data

    def _convert_xlrd_cell(
        self, cell: Any, row_idx: int, col_idx: int, worksheet: Any
    ) -> CellData | None:
        """Convert xlrd cell to CellData.

        Args:
            cell: xlrd cell object
            row_idx: Row index
            col_idx: Column index
            worksheet: xlrd worksheet

        Returns:
            CellData object or None if cell should be skipped
        """
        try:
            import xlrd

            # Get cell value based on cell type
            value = cell.value
            if cell.ctype == xlrd.XL_CELL_DATE:
                # Convert Excel date to Python datetime
                date_tuple = xlrd.xldate_as_tuple(cell.value, worksheet.book.datemode)
                value = datetime(*date_tuple) if any(date_tuple) else cell.value
            elif cell.ctype == xlrd.XL_CELL_NUMBER:
                # Convert to Decimal for precision
                with contextlib.suppress(Exception):
                    value = Decimal(str(cell.value))
            elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                value = bool(cell.value)

            data_type = self._get_data_type(value)

            # Get formatting information (limited in xlrd)
            format_info = {}
            indentation_level = 0
            alignment = None

            try:
                if hasattr(self._workbook, "format_map"):
                    cell_xf = (
                        self._workbook.xf_list[cell.xf_index]
                        if cell.xf_index < len(self._workbook.xf_list)
                        else None
                    )
                    if cell_xf:
                        # Extract font information
                        font = (
                            self._workbook.font_list[cell_xf.font_index]
                            if cell_xf.font_index < len(self._workbook.font_list)
                            else None
                        )
                        if font:
                            format_info["is_bold"] = font.bold
                            format_info["is_italic"] = font.italic
                            format_info["font_size"] = font.height / 20.0  # Convert from twips

                        # Try to extract alignment/indentation
                        if hasattr(cell_xf, "alignment"):
                            align_obj = cell_xf.alignment
                            if hasattr(align_obj, "indent_level"):
                                indentation_level = int(align_obj.indent_level)
                            if hasattr(align_obj, "hor_align"):
                                # Map xlrd alignment constants to strings
                                align_map = {0: "left", 1: "center", 2: "right"}
                                alignment = align_map.get(align_obj.hor_align, None)
            except Exception:
                pass  # Formatting extraction failed, continue without it

            return CellData(
                value=value,
                formatted_value=str(value) if value is not None else None,
                data_type=data_type,
                is_bold=format_info.get("is_bold", False),
                is_italic=format_info.get("is_italic", False),
                font_size=format_info.get("font_size"),
                has_formula=cell.ctype == xlrd.XL_CELL_FORMULA,
                indentation_level=indentation_level,
                alignment=alignment,
                # Border information not available in xlrd
                border_top=None,
                border_bottom=None,
                border_left=None,
                border_right=None,
                row=row_idx,
                column=col_idx,
            )

        except Exception as e:
            self.logger.warning(f"Failed to convert xlrd cell at {row_idx},{col_idx}: {e}")
            return None

    def _get_data_type(self, value: Any) -> str:
        """Determine data type from cell value."""
        if value is None:
            return "empty"
        elif isinstance(value, bool):
            return "boolean"
        elif isinstance(value, int | float | Decimal):
            return "number"
        elif isinstance(value, datetime):
            return "datetime"
        else:
            return "string"

    def _get_color_hex(self, color: Any) -> str | None:
        """Extract hex color from openpyxl color object."""
        try:
            if color and hasattr(color, "rgb") and color.rgb:
                return f"#{color.rgb[2:8]}"  # Remove alpha channel
        except Exception:
            pass
        return None

    def _get_fill_color_hex(self, fill: Any) -> str | None:
        """Extract background color from openpyxl fill object."""
        try:
            if fill and hasattr(fill, "start_color") and fill.start_color.rgb:
                return f"#{fill.start_color.rgb[2:8]}"  # Remove alpha channel
        except Exception:
            pass
        return None

    def _get_border_style(self, border_side: Any) -> str | None:
        """Extract border style from openpyxl border side."""
        try:
            if not border_side or not border_side.style:
                return None

            # Map openpyxl border styles to our simplified format
            style_mapping = {
                "thin": "thin",
                "medium": "medium",
                "thick": "thick",
                "hair": "thin",
                "dotted": "thin",
                "dashed": "thin",
                "dashDot": "thin",
                "dashDotDot": "thin",
                "double": "thick",
                "slantDashDot": "thin",
                "mediumDashed": "medium",
                "mediumDashDot": "medium",
                "mediumDashDotDot": "medium",
            }

            return style_mapping.get(border_side.style, "thin")
        except Exception:
            pass
        return None

    def _extract_openpyxl_metadata(self) -> dict[str, Any]:
        """Extract metadata from openpyxl workbook."""
        metadata = {}
        try:
            if hasattr(self._workbook, "properties"):
                props = self._workbook.properties
                metadata.update(
                    {
                        "title": props.title,
                        "creator": props.creator,
                        "created": props.created,
                        "modified": props.modified,
                        "last_modified_by": props.lastModifiedBy,
                        "application": getattr(props, "application", None),
                        "version": getattr(props, "appVersion", None),
                    }
                )
        except Exception as e:
            self.logger.debug(f"Failed to extract openpyxl metadata: {e}")

        return metadata

    def _extract_xlrd_metadata(self) -> dict[str, Any]:
        """Extract metadata from xlrd workbook."""
        metadata = {}
        try:
            metadata.update(
                {
                    "sheet_count": self._workbook.nsheets,
                    "date_mode": self._workbook.datemode,
                    "codepage": getattr(self._workbook, "codepage", None),
                }
            )
        except Exception as e:
            self.logger.debug(f"Failed to extract xlrd metadata: {e}")

        return metadata

    def get_excel_metadata(self) -> ExcelMetadata | None:
        """Get Excel-specific metadata (ListObjects, named ranges, print areas).

        Returns:
            ExcelMetadata object if available, None if file hasn't been read yet
        """
        return self._excel_metadata

    def get_detection_hints(self) -> list[dict[str, Any]]:
        """Get table detection hints from Excel metadata.

        Returns:
            List of detection hints with range and confidence information
        """
        if not self._excel_metadata:
            return []

        return self._metadata_extractor.convert_to_detection_hints(self._excel_metadata)
